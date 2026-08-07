#!/usr/bin/env python3
from __future__ import annotations
import io,json,re,sys
from pathlib import Path
from urllib.parse import urljoin
from openpyxl import load_workbook
ROOT=Path(__file__).resolve().parents[1]; sys.path.insert(0,str(ROOT/'scripts'))
import update_nikkei225_supply_demand as u
OUT=ROOT/'data/nikkei225-participant-file-inspection.json'; BASE='https://www.jpx.co.jp'

def latest_volume():
    m=u.get(BASE+'/automation/markets/derivatives/participant-volume/json/participant-volume_monthlylist.json').json()['TableDatas'][0]['Month']
    d=u.get(BASE+f'/automation/markets/derivatives/participant-volume/json/participant_volume_{m}.json').json()['TableDatas'][0]
    return d['TradeDate'],urljoin(BASE,d['WholeDay'])
def latest_oi():
    y=u.get(BASE+'/automation/markets/derivatives/open-interest/json/open_interest_yearlist.json').json()['TableDatas'][0]
    d=u.get(urljoin(BASE,y['Jsonfile'])).json()['TableDatas'][0]
    return d['TradeDate'],urljoin(BASE,d['IndexFutures'])
def inspect(url):
    wb=load_workbook(io.BytesIO(u.get(url).content),read_only=True,data_only=True)
    sheets=[]
    for ws in wb.worksheets:
        rows=list(ws.iter_rows(values_only=True)); out=[]
        for i,row in enumerate(rows):
            text=' | '.join(u.txt(x) for x in row if x not in (None,''))
            if re.search(r'日経\s*225|Nikkei\s*225',text,re.I):
                for j in range(max(0,i-5),min(len(rows),i+35)):
                    cells=[[c,u.txt(v)] for c,v in enumerate(rows[j]) if v not in (None,'')]
                    out.append({'row':j+1,'cells':cells})
        if out:
            seen=set(); uniq=[]
            for x in out:
                if x['row'] not in seen:seen.add(x['row']);uniq.append(x)
            sheets.append({'title':ws.title,'rows':uniq[:180]})
    return sheets
vd,vu=latest_volume(); od,ou=latest_oi()
payload={'generatedAt':u.now(),'participantVolume':{'tradeDate':vd,'url':vu,'sheets':inspect(vu)},'participantOpenInterest':{'tradeDate':od,'url':ou,'sheets':inspect(ou)}}
OUT.write_text(json.dumps(payload,ensure_ascii=False,indent=2)+'\n',encoding='utf-8'); print(OUT)
