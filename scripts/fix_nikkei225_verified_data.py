#!/usr/bin/env python3
"""Final verified-data fixes for current JPX Nikkei 225 layouts."""
from __future__ import annotations
import io, json, re
from pathlib import Path
from typing import Any
from openpyxl import load_workbook
import update_nikkei225_supply_demand as u

ROOT=Path(__file__).resolve().parents[1]
OUT=ROOT/'data/nikkei225-supply-demand.json'

def ni(v:Any):
    x=u.n(v)
    return int(round(x)) if x is not None else None

def latest_oi_url():
    xs=[url for url,_ in u.links(u.URLS['futures']) if re.search(r'open_interest\.xlsx(?:\?|$)',url,re.I)]
    return xs[0] if xs else None

def parse_oi(url:str):
    wb=load_workbook(io.BytesIO(u.get(url).content),read_only=True,data_only=True)
    targets={'日経225':'large','日経225mini':'mini'}
    found={}
    for ws in wb.worksheets:
        state={0:None,6:None}
        accum={}
        for raw in ws.iter_rows(values_only=True):
            row=list(raw)
            for base in (0,6):
                if base>=len(row): continue
                c0=u.txt(row[base]); c1=u.txt(row[base+1]) if base+1<len(row) else ''
                # First contract row: product name in product column.
                if c0 in targets:
                    key=targets[c0]; state[base]=key
                    accum[(base,key)]={'volume':0,'openInterest':0,'openInterestChange':0,'previousOpenInterest':0}
                    for name,idx in (('volume',base+2),('openInterest',base+3),('openInterestChange',base+4),('previousOpenInterest',base+5)):
                        if idx<len(row) and ni(row[idx]) is not None: accum[(base,key)][name]+=ni(row[idx])
                    continue
                key=state.get(base)
                if not key: continue
                # Current JPX sheet keeps product cell blank on continuation and
                # total rows. Contract/"合計" is therefore in base+1.
                if not c0 and u.norm(c1) in {'合計','total'}:
                    vals={
                        'volume':ni(row[base+2]) if base+2<len(row) else None,
                        'openInterest':ni(row[base+3]) if base+3<len(row) else None,
                        'openInterestChange':ni(row[base+4]) if base+4<len(row) else None,
                        'previousOpenInterest':ni(row[base+5]) if base+5<len(row) else None,
                    }
                    a=accum[(base,key)]
                    for k in vals:
                        if vals[k] is None: vals[k]=a[k]
                    found[key]=vals; state[base]=None; continue
                # Some workbook versions put 合計 one cell left.
                if u.norm(c0) in {'合計','total'}:
                    vals={
                        'volume':ni(row[base+1]) if base+1<len(row) else None,
                        'openInterest':ni(row[base+2]) if base+2<len(row) else None,
                        'openInterestChange':ni(row[base+3]) if base+3<len(row) else None,
                        'previousOpenInterest':ni(row[base+4]) if base+4<len(row) else None,
                    }
                    a=accum[(base,key)]
                    for k in vals:
                        if vals[k] is None: vals[k]=a[k]
                    found[key]=vals; state[base]=None; continue
                # Continuation month row: blank product, contract in base+1.
                if not c0 and re.search(r'20\d{2}年\d{2}月限',c1):
                    a=accum[(base,key)]
                    for name,idx in (('volume',base+2),('openInterest',base+3),('openInterestChange',base+4),('previousOpenInterest',base+5)):
                        if idx<len(row) and ni(row[idx]) is not None: a[name]+=ni(row[idx])
                    continue
                # Stop state only when another explicit product name appears.
                if c0 and not re.search(r'20\d{2}年\d{2}月限',c0): state[base]=None
    return found

def main():
    d=json.loads(OUT.read_text(encoding='utf-8'))
    f=d.setdefault('futures',{})
    try:
        url=latest_oi_url()
        if not url: raise ValueError('open_interest.xlsx not found')
        x=parse_oi(url); large=x.get('large'); mini=x.get('mini')
        if not large or large.get('openInterest') is None: raise ValueError(f'large total not found: {x}')
        f['volume']=large.get('volume'); f['openInterest']=large.get('openInterest'); f['openInterestChange']=large.get('openInterestChange')
        f['previousOpenInterest']=large.get('previousOpenInterest')
        if mini:
            f['miniVolume']=mini.get('volume'); f['miniOpenInterest']=mini.get('openInterest'); f['miniOpenInterestChange']=mini.get('openInterestChange'); f['miniPreviousOpenInterest']=mini.get('previousOpenInterest')
        m=re.search(r'(20\d{6})',url)
        if m:
            from datetime import datetime
            f['asOfDate']=datetime.strptime(m.group(1),'%Y%m%d').date().isoformat()
        f['oiSourceUrl']=url; f['status']='verified' if u.n(f.get('price')) is not None else 'partial'; f.pop('error',None); f['fetchedAt']=u.now()
    except Exception as exc:
        f['status']='partial' if u.n(f.get('price')) is not None else 'unavailable'; f['error']=f'JPX建玉取得失敗: {type(exc).__name__}: {exc}'; f['fetchedAt']=u.now()
    # Never label a five-observation mean as a 20-day average.
    s=d.get('shortSelling') or {}
    count=int(s.get('sampleCount') or 0)
    if count<20: s['avg20']=None
    if count<5: s['avg5']=None
    d['shortSelling']=s
    d['assessment']=u.assessment(d.get('futures') or {},d.get('arbitrage') or {},d.get('options') or {},d.get('foreignInvestors') or {})
    keys=('spot','futures','sessions','arbitrage','options','participantFlow','foreignInvestors','participantOpenInterest','shortSelling','margin')
    statuses={k:(d.get(k) or {}).get('status','unavailable') for k in keys}; connected=sum(v in {'verified','calculated'} for v in statuses.values())
    d['sourceStatus']=f'{connected}/10項目連携（基準日を個別表示）'; d.setdefault('diagnostics',{})['statuses']=statuses; d['diagnostics']['oiParser']='verified current JPX horizontal-block layout'; d['generatedAt']=u.now()
    OUT.write_text(json.dumps(d,ensure_ascii=False,indent=2)+'\n',encoding='utf-8')
    print(json.dumps({'large':f.get('openInterest'),'largeChange':f.get('openInterestChange'),'largeVolume':f.get('volume'),'mini':f.get('miniOpenInterest'),'miniChange':f.get('miniOpenInterestChange'),'miniVolume':f.get('miniVolume'),'status':f.get('status')},ensure_ascii=False))
if __name__=='__main__': main()
