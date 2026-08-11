#!/usr/bin/env python3
"""Build the gold supply-demand JSON from public, verifiable sources.

Principles:
- Never fabricate a missing value.
- Keep each data set's own as-of date.
- Preserve the last verified slow-moving value when a source is temporarily
  unavailable, but mark it stale.
- Separate short-term financial demand from structural physical demand.
"""
from __future__ import annotations

import csv
import io
import json
import math
import re
import time
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "gold-supply-demand.json"
MARKET = ROOT / "data" / "market" / "latest.json"
JST = timezone(timedelta(hours=9))
NY = ZoneInfo("America/New_York")
UA = "Mozilla/5.0 (compatible; ChatGPT-Market-Report/1.0; +https://github.com/matrixdiamond512-cell/Chat-GPT-Market-Report)"

CME_PDF = "https://www.cmegroup.com/daily_bulletin/current/Section02B_Summary_Volume_And_Open_Interest_Metals_Futures_And_Options.pdf"
CFTC_API = "https://publicreporting.cftc.gov/resource/72hh-3qpy.json"
GLD_XLSX = "https://api.spdrgoldshares.com/api/v1/historical-archive?exchange=NYSE&lang=en&product=gld"
IAU_URL = "https://www.ishares.com/us/products/239561/ishares-gold-trust-fund"
FRED_CSV = "https://fred.stlouisfed.org/graph/fredgraph.csv?id={series}"
YAHOO_CHART = "https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
WGC_PREMIUM = "https://www.gold.org/goldhub/data/gold-premium"
WGC_ETF = "https://www.gold.org/goldhub/data/gold-etfs-holdings-and-flows"
WGC_CENTRAL = "https://www.gold.org/goldhub/gold-focus/2026/07/central-bank-gold-statistics-central-banks-remain-committed-gold"


def now_iso() -> str:
    return datetime.now(JST).replace(microsecond=0).isoformat()


def get_json(path: Path, default: Any) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def num(v: Any) -> float | None:
    try:
        x = float(str(v).replace(",", "").strip())
        return x if math.isfinite(x) else None
    except Exception:
        return None


def i_num(v: Any) -> int | None:
    x = num(v)
    return int(round(x)) if x is not None else None


def parse_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v or "").strip()
    for f in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d-%b-%Y", "%b %d, %Y", "%B %d, %Y"):
        try:
            return datetime.strptime(s, f).date()
        except ValueError:
            pass
    return None


def request(url: str, *, params: dict[str, Any] | None = None, timeout: int | tuple[int, ...] = 30) -> requests.Response:
    timeouts = timeout if isinstance(timeout, tuple) else (timeout,)
    last_error: Exception | None = None
    for attempt, seconds in enumerate(timeouts, start=1):
        try:
            response = requests.get(url, params=params, timeout=seconds, headers={"User-Agent": UA, "Accept": "*/*"})
            response.raise_for_status()
            return response
        except requests.RequestException as exc:
            last_error = exc
            if attempt < len(timeouts):
                time.sleep(min(2 ** (attempt - 1), 8))
    raise RuntimeError(f"HTTP取得失敗（{len(timeouts)}回試行）: {last_error}")


def stale_copy(previous: dict[str, Any] | None, reason: str) -> dict[str, Any]:
    if not isinstance(previous, dict) or not any(v not in (None, "", []) for k, v in previous.items() if k not in {"status", "error"}):
        return {"status": "unavailable", "error": reason}
    out = dict(previous)
    out["status"] = "preserved_after_fetch_error"
    out["error"] = reason
    out["lastAttemptAt"] = now_iso()
    return out


def verified(payload: dict[str, Any] | None) -> bool:
    return isinstance(payload, dict) and payload.get("status") == "verified"


def fetch_yahoo_daily(symbol: str) -> list[tuple[date, float]]:
    try:
        r = request(YAHOO_CHART.format(symbol=requests.utils.quote(symbol, safe="")), params={"range": "15d", "interval": "1d", "events": "history"})
        result = (r.json().get("chart", {}).get("result") or [None])[0] or {}
        stamps = result.get("timestamp") or []
        closes = (((result.get("indicators") or {}).get("quote") or [{}])[0].get("close") or [])
        out: list[tuple[date, float]] = []
        for ts, close in zip(stamps, closes):
            x = num(close)
            if x is None:
                continue
            d = datetime.fromtimestamp(int(ts), timezone.utc).astimezone(NY).date()
            out.append((d, x))
        return out
    except Exception:
        return []


def fetch_comex(previous: dict[str, Any] | None) -> dict[str, Any]:
    try:
        pdf = request(CME_PDF, timeout=(20, 30, 45)).content
        reader = PdfReader(io.BytesIO(pdf))
        text = "\n".join((p.extract_text() or "") for p in reader.pages)
        dm = re.search(r"(?:Mon|Tue|Wed|Thu|Fri),\s+([A-Z][a-z]{2})\s+(\d{1,2}),\s+(\d{4})", text)
        if not dm:
            raise ValueError("CME bulletin date not found")
        asof = datetime.strptime(f"{dm.group(1)} {dm.group(2)} {dm.group(3)}", "%b %d %Y").date()
        # GC row columns: Globex volume, PNT volume, total volume, OI, OI sign/change, year-ago volume/OI.
        m = re.search(r"GC\s+COMEX GOLD FUTURES\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([+-])\s*([\d,]+)", text)
        if not m:
            raise ValueError("GC COMEX GOLD FUTURES row not found")
        volume = i_num(m.group(3))
        oi = i_num(m.group(4))
        oi_change = i_num(m.group(6))
        if m.group(5) == "-" and oi_change is not None:
            oi_change *= -1
        price = price_prev = price_change_pct = None
        daily = fetch_yahoo_daily("GC=F")
        for idx, (d, p) in enumerate(daily):
            if d == asof:
                price = p
                if idx > 0:
                    price_prev = daily[idx - 1][1]
                break
        if price is not None and price_prev not in (None, 0):
            price_change_pct = (price / price_prev - 1.0) * 100.0
        interpretation = None
        if price_change_pct is not None and oi_change is not None:
            if price_change_pct > 0 and oi_change > 0:
                interpretation = "新規ロング流入の可能性"
            elif price_change_pct > 0 and oi_change < 0:
                interpretation = "ショートカバー中心の可能性"
            elif price_change_pct < 0 and oi_change > 0:
                interpretation = "新規ショート流入の可能性"
            elif price_change_pct < 0 and oi_change < 0:
                interpretation = "ロング清算の可能性"
            else:
                interpretation = "方向感限定"
        return {
            "status": "verified",
            "asOfDate": asof.isoformat(),
            "volume": volume,
            "openInterest": oi,
            "openInterestChange": oi_change,
            "alignedPrice": price,
            "alignedPriceChangePercent": price_change_pct,
            "interpretation": interpretation,
            "sourceName": "CME Group Daily Bulletin",
            "sourceUrl": CME_PDF,
            "fetchedAt": now_iso(),
        }
    except Exception as exc:
        return stale_copy(previous, f"CME取得失敗: {type(exc).__name__}: {exc}")


def fetch_cftc(previous: dict[str, Any] | None) -> dict[str, Any]:
    try:
        fields = ",".join([
            "report_date_as_yyyy_mm_dd", "open_interest_all", "m_money_positions_long_all", "m_money_positions_short_all"
        ])
        params = {
            "$select": fields,
            "$where": "cftc_contract_market_code='088691'",
            "$order": "report_date_as_yyyy_mm_dd DESC",
            "$limit": "2",
        }
        rows = request(CFTC_API, params=params).json()
        if not isinstance(rows, list) or not rows:
            raise ValueError("CFTC returned no GOLD rows")
        cur = rows[0]
        prev = rows[1] if len(rows) > 1 else {}
        long_ = i_num(cur.get("m_money_positions_long_all"))
        short = i_num(cur.get("m_money_positions_short_all"))
        oi = i_num(cur.get("open_interest_all"))
        prev_long = i_num(prev.get("m_money_positions_long_all"))
        prev_short = i_num(prev.get("m_money_positions_short_all"))
        net = long_ - short if long_ is not None and short is not None else None
        prev_net = prev_long - prev_short if prev_long is not None and prev_short is not None else None
        net_change = net - prev_net if net is not None and prev_net is not None else None
        judgement = "中立"
        if net_change is not None:
            judgement = "投機筋の買い増加" if net_change > 0 else "投機筋の売り増加" if net_change < 0 else "投機筋は横ばい"
        asof_raw = str(cur.get("report_date_as_yyyy_mm_dd") or "")[:10]
        return {
            "status": "verified",
            "asOfDate": asof_raw,
            "managedMoneyLong": long_,
            "managedMoneyShort": short,
            "managedMoneyNet": net,
            "managedMoneyNetChange": net_change,
            "openInterest": oi,
            "judgement": judgement,
            "sourceName": "CFTC Disaggregated COT - Futures Only",
            "sourceUrl": "https://publicreporting.cftc.gov/Commitments-of-Traders/Disaggregated-Futures-Only/72hh-3qpy",
            "fetchedAt": now_iso(),
        }
    except Exception as exc:
        return stale_copy(previous, f"CFTC取得失敗: {type(exc).__name__}: {exc}")


def _find_excel_table(content: bytes, value_header_keywords: tuple[str, ...]) -> list[tuple[date, float]]:
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    candidates: list[tuple[date, float]] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        header_i = date_col = value_col = None
        for ri, row in enumerate(rows[:80]):
            texts = [str(x or "").strip().lower() for x in row]
            for ci, t in enumerate(texts):
                if date_col is None and (t == "date" or "date" in t):
                    date_col = ci
                if value_col is None and any(k.lower() in t for k in value_header_keywords):
                    value_col = ci
            if date_col is not None and value_col is not None:
                header_i = ri
                break
        if header_i is None or date_col is None or value_col is None:
            continue
        for row in rows[header_i + 1 :]:
            if date_col >= len(row) or value_col >= len(row):
                continue
            d = parse_date(row[date_col])
            v = num(row[value_col])
            if d and v is not None:
                candidates.append((d, v))
    candidates.sort(key=lambda x: x[0])
    return candidates


def fetch_gld(previous: dict[str, Any] | None) -> dict[str, Any]:
    try:
        content = request(GLD_XLSX, timeout=45).content
        rows = _find_excel_table(content, ("tonnes", "tonnes of gold", "gold holdings"))
        if not rows:
            raise ValueError("GLD tonnes column not found in historical archive")
        d, tonnes = rows[-1]
        change = tonnes - rows[-2][1] if len(rows) > 1 else None
        return {
            "status": "verified",
            "asOfDate": d.isoformat(),
            "tonnes": tonnes,
            "changeTonnes": change,
            "sourceName": "SPDR Gold Shares Historical Archive",
            "sourceUrl": "https://www.spdrgoldshares.com/usa/gld/",
            "fetchedAt": now_iso(),
        }
    except Exception as exc:
        return stale_copy(previous, f"GLD取得失敗: {type(exc).__name__}: {exc}")


def fetch_iau(previous: dict[str, Any] | None) -> dict[str, Any]:
    try:
        html = request(IAU_URL, timeout=30).text
        text = BeautifulSoup(html, "html.parser").get_text(" ", strip=True)
        m = re.search(r"Tonnes in Trust\s*([\d,]+(?:\.\d+)?)\s*as of\s*([A-Z][a-z]{2}\s+\d{1,2},\s+\d{4})", text, re.I)
        if not m:
            # Some page variants place the date before the number in generated markup.
            m2 = re.search(r"Tonnes in Trust.{0,160}?([\d,]+(?:\.\d+)?).{0,80}?([A-Z][a-z]{2}\s+\d{1,2},\s+\d{4})", text, re.I)
            if not m2:
                raise ValueError("IAU Tonnes in Trust not found")
            m = m2
        tonnes = num(m.group(1))
        d = parse_date(m.group(2))
        if tonnes is None or d is None:
            raise ValueError("IAU tonnes/date parse failed")
        change = None
        if isinstance(previous, dict) and parse_date(previous.get("asOfDate")) and num(previous.get("tonnes")) is not None:
            pd = parse_date(previous.get("asOfDate"))
            pv = num(previous.get("tonnes"))
            if pd and pd < d and pv is not None:
                change = tonnes - pv
            elif pd == d:
                change = num(previous.get("changeTonnes"))
        return {
            "status": "verified",
            "asOfDate": d.isoformat(),
            "tonnes": tonnes,
            "changeTonnes": change,
            "sourceName": "iShares Gold Trust",
            "sourceUrl": IAU_URL,
            "fetchedAt": now_iso(),
        }
    except Exception as exc:
        return stale_copy(previous, f"IAU取得失敗: {type(exc).__name__}: {exc}")


def fetch_fred(series: str, previous: dict[str, Any] | None, label: str) -> dict[str, Any]:
    try:
        text = request(FRED_CSV.format(series=series)).text
        rows = []
        for row in csv.DictReader(io.StringIO(text)):
            d = parse_date(row.get("DATE") or row.get("observation_date"))
            v = num(row.get(series))
            if d and v is not None:
                rows.append((d, v))
        if not rows:
            raise ValueError(f"FRED {series} returned no observations")
        d, value = rows[-1]
        change = value - rows[-2][1] if len(rows) > 1 else None
        return {
            "status": "verified",
            "asOfDate": d.isoformat(),
            "value": value,
            "change": change,
            "sourceName": f"FRED {series}",
            "sourceUrl": f"https://fred.stlouisfed.org/series/{series}",
            "label": label,
            "fetchedAt": now_iso(),
        }
    except Exception as exc:
        return stale_copy(previous, f"FRED {series}取得失敗: {type(exc).__name__}: {exc}")


def fetch_wgc_premiums(previous: dict[str, Any] | None) -> dict[str, Any]:
    """Try the official downloadable workbook; keep stale data if login blocks it."""
    prev = previous if isinstance(previous, dict) else {}
    try:
        html = request(WGC_PREMIUM).text
        soup = BeautifulSoup(html, "html.parser")
        links = []
        for a in soup.find_all("a", href=True):
            text = a.get_text(" ", strip=True).lower()
            href = str(a.get("href"))
            if "premium" in text and ("xlsx" in text or href.lower().endswith((".xlsx", ".xls"))):
                links.append(requests.compat.urljoin(WGC_PREMIUM, href))
        if not links:
            raise ValueError("WGC premium workbook requires sign-in or no direct XLSX link was exposed")
        content = request(links[0], timeout=45).content
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        data: list[tuple[date, float | None, float | None]] = []
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            h = None
            for ri, row in enumerate(rows[:60]):
                tx = [str(v or "").lower() for v in row]
                dc = next((i for i, t in enumerate(tx) if "date" in t), None)
                cc = next((i for i, t in enumerate(tx) if "china" in t), None)
                ic = next((i for i, t in enumerate(tx) if "india" in t), None)
                if dc is not None and cc is not None and ic is not None:
                    h = (ri, dc, cc, ic)
                    break
            if not h:
                continue
            ri, dc, cc, ic = h
            for row in rows[ri + 1 :]:
                d = parse_date(row[dc] if dc < len(row) else None)
                if d:
                    data.append((d, num(row[cc] if cc < len(row) else None), num(row[ic] if ic < len(row) else None)))
        data = [x for x in data if x[1] is not None or x[2] is not None]
        data.sort(key=lambda x: x[0])
        if not data:
            raise ValueError("WGC premium workbook columns not parsed")
        d, china, india = data[-1]
        prev_china = data[-2][1] if len(data) > 1 else None
        prev_india = data[-2][2] if len(data) > 1 else None
        return {
            "status": "verified",
            "asOfDate": d.isoformat(),
            "china": {"premiumUsdOz": china, "change": china - prev_china if china is not None and prev_china is not None else None},
            "india": {"premiumUsdOz": india, "change": india - prev_india if india is not None and prev_india is not None else None},
            "sourceName": "World Gold Council Local gold price premium/discount",
            "sourceUrl": WGC_PREMIUM,
            "fetchedAt": now_iso(),
        }
    except Exception as exc:
        return {
            "status": "unavailable",
            "frequency": "weekly",
            "sourceName": "World Gold Council Local gold price premium/discount",
            "sourceUrl": WGC_PREMIUM,
            "fetchedAt": now_iso(),
            "error": f"WGC現物プレミアム取得不能: {type(exc).__name__}: {exc}",
            "note": "一般公開経路を確認できないため旧値を最新値として表示しません。",
        }


def assessment(data: dict[str, Any]) -> dict[str, Any]:
    components: list[tuple[str, float]] = []
    comex = data.get("comex") or {}
    cftc = data.get("cftc") or {}
    etf = data.get("etf") or {}
    env = data.get("environment") or {}

    if verified(comex) and comex.get("interpretation"):
        score = {
            "新規ロング流入の可能性": 1.0,
            "ショートカバー中心の可能性": 0.4,
            "新規ショート流入の可能性": -1.0,
            "ロング清算の可能性": -0.4,
        }.get(comex.get("interpretation"), 0.0)
        components.append(("COMEX", score))
    if verified(cftc) and num(cftc.get("managedMoneyNetChange")) is not None:
        components.append(("CFTC", 1.0 if num(cftc.get("managedMoneyNetChange")) > 0 else -1.0 if num(cftc.get("managedMoneyNetChange")) < 0 else 0.0))
    changes = []
    for k in ("gld", "iau"):
        x = etf.get(k) or {}
        if verified(x) and num(x.get("changeTonnes")) is not None:
            changes.append(num(x.get("changeTonnes")))
    if changes:
        s = sum(changes)
        components.append(("ETF", 1.0 if s > 0 else -1.0 if s < 0 else 0.0))
    ry = env.get("realYield10y") or {}
    if verified(ry) and num(ry.get("change")) is not None:
        c = num(ry.get("change"))
        components.append(("実質金利", 1.0 if c < 0 else -1.0 if c > 0 else 0.0))
    db = env.get("dollarBroad") or {}
    if verified(db) and num(db.get("change")) is not None:
        c = num(db.get("change"))
        components.append(("ドル", 1.0 if c < 0 else -1.0 if c > 0 else 0.0))

    if len(components) >= 3:
        avg = sum(v for _, v in components) / len(components)
        score = round(max(0, min(100, 50 + avg * 35)))
        if avg >= 0.45:
            short = "買い優勢"
        elif avg <= -0.45:
            short = "売り優勢"
        else:
            short = "中立〜やや買い" if avg > 0.1 else "中立〜やや売り" if avg < -0.1 else "中立"
    else:
        score = None
        short = "判定待ち"

    structural_parts: list[float] = []
    cb = data.get("centralBank") or {}
    physical = data.get("physical") or {}
    if cb.get("status") in {"verified", "stale"} and num(cb.get("netPurchasesTonnes")) is not None:
        x = num(cb.get("netPurchasesTonnes"))
        structural_parts.append(1.0 if x > 0 else -1.0 if x < 0 else 0.0)
    for k in ("china", "india"):
        x = (physical.get(k) or {}) if isinstance(physical, dict) else {}
        if x.get("status") == "verified" and num(x.get("premiumUsdOz")) is not None:
            p = num(x.get("premiumUsdOz"))
            structural_parts.append(1.0 if p > 0 else -1.0 if p < 0 else 0.0)
    if len(structural_parts) >= 2:
        a = sum(structural_parts) / len(structural_parts)
        structural = "強い" if a >= 0.4 else "弱い" if a <= -0.4 else "中立"
    elif structural_parts and structural_parts[0] > 0:
        structural = "買い需要あり（確認項目限定）"
    else:
        structural = "判定待ち"

    return {"shortTerm": short, "structural": structural, "score": score, "components": [{"name": n, "score": v} for n, v in components]}


def build_summary(data: dict[str, Any], market: dict[str, Any]) -> list[str]:
    out: list[str] = []
    gold = ((market.get("markets") or {}).get("gold") or {})
    if gold.get("verificationStatus") == "verified" and num(gold.get("changePercent")) is not None:
        ch = num(gold.get("changePercent"))
        out.append(f"金先物は前回比{ch:+.2f}%です。COMEXの出来高・建玉は別の基準日で表示し、日付を合わせられた場合だけ価格×建玉判定を行います。")
    c = data.get("comex") or {}
    if verified(c):
        out.append(f"COMEX GCは{c.get('asOfDate')}基準で出来高{int(c.get('volume') or 0):,}枚、建玉{int(c.get('openInterest') or 0):,}枚、建玉前日比{int(c.get('openInterestChange') or 0):+,}枚です。")
    else:
        out.append("COMEX出来高・建玉は取得待ちです。取得不能時は推測値を表示しません。")
    cftc = data.get("cftc") or {}
    if cftc.get("status") in {"verified", "stale"} and num(cftc.get("managedMoneyNet")) is not None:
        net = i_num(cftc.get("managedMoneyNet")) or 0
        change = i_num(cftc.get("managedMoneyNetChange"))
        out.append(f"CFTC Managed Moneyのネットは{net:,}枚" + (f"、前週比{change:+,}枚" if change is not None else "") + "です。週次データとして短期需給に反映します。")
    etf = data.get("etf") or {}
    etf_bits = []
    for name, k in (("GLD", "gld"), ("IAU", "iau")):
        x = etf.get(k) or {}
        if x.get("status") in {"verified", "stale"} and num(x.get("tonnes")) is not None:
            etf_bits.append(f"{name} {num(x.get('tonnes')):.2f}t")
    out.append(("主要金ETF保有量は " + " / ".join(etf_bits) + "。金融需要の実流入・流出を先物と分けて確認します。") if etf_bits else "GLD・IAU保有量は取得待ちです。")
    p = data.get("physical") or {}
    if p.get("status") == "verified":
        out.append("中国・インドの現物プレミアムを使い、金融市場主導か現物需要主導かを切り分けます。")
    else:
        out.append("中国・インドの現物プレミアムはWGCの公開更新を確認できる場合だけ表示します。ログイン等で取得できない場合は前回値を捏造せず保留します。")
    cb = data.get("centralBank") or {}
    if num(cb.get("netPurchasesTonnes")) is not None:
        out.append(f"中央銀行需要は{cb.get('period') or '最新月'}の純購入{num(cb.get('netPurchasesTonnes')):+.0f}tを構造的需給として扱い、日次シグナルとは分離します。")
    return out[:5]


def main() -> int:
    previous = get_json(OUT, {})
    market = get_json(MARKET, {})
    prev_etf = previous.get("etf") if isinstance(previous, dict) else {}
    prev_env = previous.get("environment") if isinstance(previous, dict) else {}

    data: dict[str, Any] = {
        "schemaVersion": "2.0.0",
        "pageTitle": "ゴールド需給分析",
        "subtitle": "COMEX先物・投機筋・ETF・中国/インド現物需要から総合判定",
        "generatedAt": now_iso(),
        "comex": fetch_comex(previous.get("comex") if isinstance(previous, dict) else None),
        "cftc": fetch_cftc(previous.get("cftc") if isinstance(previous, dict) else None),
        "etf": {
            "gld": fetch_gld((prev_etf or {}).get("gld") if isinstance(prev_etf, dict) else None),
            "iau": fetch_iau((prev_etf or {}).get("iau") if isinstance(prev_etf, dict) else None),
            "global": stale_copy((prev_etf or {}).get("global") if isinstance(prev_etf, dict) else None, "WGC世界ETFは週次/月次の公開値を自動取得できる場合のみ更新"),
        },
        "physical": fetch_wgc_premiums(previous.get("physical") if isinstance(previous, dict) else None),
        "curve": {
            "status": "unavailable",
            "implementationStatus": "not_implemented",
            "frequency": "daily",
            "error": "未実装・安定データソース確認中",
            "note": "安定した無料公開データ連携を確認後に実装します。",
        },
        "centralBank": previous.get("centralBank") if isinstance(previous.get("centralBank"), dict) else {
            "status": "stale",
            "period": "2026-05",
            "netPurchasesTonnes": 41,
            "sourceName": "World Gold Council central bank gold statistics",
            "sourceUrl": WGC_CENTRAL,
            "note": "月次データ。2026-07-02公表記事の2026年5月純購入量。",
        },
        "environment": {
            "realYield10y": fetch_fred("DFII10", (prev_env or {}).get("realYield10y") if isinstance(prev_env, dict) else None, "米10年実質金利"),
            "dollarBroad": fetch_fred("DTWEXBGS", (prev_env or {}).get("dollarBroad") if isinstance(prev_env, dict) else None, "米ドル実効指数（Broad）"),
            "usdjpy": {},
        },
    }
    data["centralBank"]["frequency"] = "monthly"
    usd = ((market.get("markets") or {}).get("usdjpy") or {})
    if usd.get("verificationStatus") == "verified" and num(usd.get("value")) is not None:
        data["environment"]["usdjpy"] = {
            "status": "verified",
            "asOfDateTime": usd.get("asOf"),
            "value": num(usd.get("value")),
            "change": num(usd.get("change")),
            "changePercent": num(usd.get("changePercent")),
            "sourceName": usd.get("sourceName"),
            "sourceUrl": usd.get("sourceUrl"),
        }
    else:
        data["environment"]["usdjpy"] = {"status": "unavailable", "error": "検証済みUSD/JPY未取得"}

    # Flatten physical child status for the page while retaining source-level status.
    p = data.get("physical") or {}
    if isinstance(p, dict) and p.get("status") in {"verified", "stale"}:
        for k in ("china", "india"):
            if isinstance(p.get(k), dict):
                p[k]["status"] = p.get("status")
                p[k]["asOfDate"] = p.get("asOfDate")

    data["assessment"] = assessment(data)
    groups = [
        ((market.get("markets") or {}).get("gold") or {}).get("verificationStatus") == "verified",
        verified(data.get("comex")),
        verified(data.get("cftc")),
        any(verified((data.get("etf") or {}).get(k)) for k in ("gld", "iau")),
        (data.get("physical") or {}).get("status") == "verified",
        (data.get("centralBank") or {}).get("status") in {"verified", "stale"},
        any(verified((data.get("environment") or {}).get(k)) for k in ("realYield10y", "dollarBroad", "usdjpy")),
    ]
    data["dataStatus"] = {"connected": sum(bool(x) for x in groups), "total": 7}
    data["aiSummary"] = build_summary(data, market)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"generatedAt": data["generatedAt"], "dataStatus": data["dataStatus"], "assessment": data["assessment"]}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
