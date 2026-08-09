#!/usr/bin/env python3
"""Parse JPX participant turnover/OI and build a same-date Nikkei 225 analysis context."""
from __future__ import annotations
import io,json,re
from datetime import datetime
from pathlib import Path
from urllib.parse import urljoin
from openpyxl import load_workbook
import update_nikkei225_supply_demand as u
import fix_nikkei225_verified_data as vf

ROOT=Path(__file__).resolve().parents[1]
OUT=ROOT/'data/nikkei225-supply-demand.json'
BASE='https://www.jpx.co.jp'

def iso_date(s:str): return datetime.strptime(str(s),'%Y%m%d').date().isoformat()
def ymd(s:str): return str(s).replace('-','')[:8]
def same_day(a,b): return str(a or '')[:10]==str(b or '')[:10]

def latest_participant_volume_file():
    ml=u.get(BASE+'/automation/markets/derivatives/participant-volume/json/participant-volume_monthlylist.json').json()
    month=ml['TableDatas'][0]['Month']
    cur=u.get(BASE+f'/automation/markets/derivatives/participant-volume/json/participant_volume_{month}.json').json()
    row=cur['TableDatas'][0]
    return row['TradeDate'],urljoin(BASE,row['WholeDay'])

def participant_volume_file_for_date(trade_date:str):
    td=ymd(trade_date); month=td[:6]
    cur=u.get(BASE+f'/automation/markets/derivatives/participant-volume/json/participant_volume_{month}.json').json()
    for row in cur.get('TableDatas',[]):
        if ymd(row.get('TradeDate'))==td and row.get('WholeDay'):
            return row['TradeDate'],urljoin(BASE,row['WholeDay'])
    raise ValueError(f'participant turnover file not found for {td}')

def parse_turnover_leaders(url:str):
    wb=load_workbook(io.BytesIO(u.get(url).content),read_only=True,data_only=True)
    rows=[]
    for ws in wb.worksheets:
        for raw in ws.iter_rows(values_only=True):
            r=list(raw)
            if len(r)<8 or u.txt(r[0])!='NK225F': continue
            contract=u.txt(r[2]); rank=u.n(r[3]); name=u.txt(r[5]); vol=u.n(r[7])
            if not contract or rank is None or not name or vol is None: continue
            rows.append({'contract':contract,'rank':int(rank),'name':name,'volume':int(round(vol))})
    if not rows: raise ValueError('NK225F turnover rows not found')
    front=rows[0]['contract']
    leaders=[{'name':x['name'],'volume':x['volume'],'rank':x['rank']} for x in rows if x['contract']==front]
    leaders.sort(key=lambda x:x['rank'])
    return front,leaders[:10]

def participant_oi_file_rows():
    yl=u.get(BASE+'/automation/markets/derivatives/open-interest/json/open_interest_yearlist.json').json()
    out=[]
    # Current and previous year are enough for latest-week and prior-week comparison,
    # including the turn of the year.
    for year in yl.get('TableDatas',[])[:2]:
        jf=year.get('Jsonfile') if isinstance(year,dict) else None
        if not jf: continue
        cur=u.get(urljoin(BASE,jf)).json()
        for row in cur.get('TableDatas',[]):
            td=ymd(row.get('TradeDate'))
            file=row.get('IndexFutures')
            if re.fullmatch(r'20\d{6}',td) and file:
                out.append((td,urljoin(BASE,file)))
    unique={td:url for td,url in out}
    return sorted(unique.items(),key=lambda x:x[0],reverse=True)

def latest_participant_oi_file():
    rows=participant_oi_file_rows()
    if not rows: raise ValueError('participant OI weekly list is empty')
    return rows[0]

def previous_participant_oi_file(trade_date:str):
    td=ymd(trade_date); rows=participant_oi_file_rows()
    older=[x for x in rows if x[0]<td]
    if not older: return None,None
    return older[0]

def parse_oi_leaders(url:str):
    wb=load_workbook(io.BytesIO(u.get(url).content),read_only=True,data_only=True)
    for ws in wb.worksheets:
        rows=list(ws.iter_rows(values_only=True)); start=None
        for i,raw in enumerate(rows):
            first=u.txt(raw[0]) if len(raw)>0 else ''
            if first=='＜日経225先物＞': start=i+1; break
        if start is None: continue
        first_contract=None; sellers=[]; buyers=[]
        for raw in rows[start:]:
            r=list(raw); first=u.txt(r[0]) if len(r)>0 else ''
            if first.startswith('＜') and first.endswith('＞'): break
            contract=u.txt(r[1]) if len(r)>1 else ''; rank=u.n(r[0]) if len(r)>0 else None
            if rank is None or not re.search(r'20\d{2}年\d{2}月限月',contract): continue
            if first_contract is None: first_contract=contract
            if contract!=first_contract: continue
            sell_name=u.txt(r[3]) if len(r)>3 else ''; sell_qty=u.n(r[4]) if len(r)>4 else None
            buy_name=u.txt(r[6]) if len(r)>6 else ''; buy_qty=u.n(r[7]) if len(r)>7 else None
            if sell_name and sell_qty is not None: sellers.append({'name':sell_name,'openInterest':int(round(sell_qty)),'rank':int(rank)})
            if buy_name and buy_qty is not None: buyers.append({'name':buy_name,'openInterest':int(round(buy_qty)),'rank':int(rank)})
        if first_contract and (sellers or buyers):
            sellers.sort(key=lambda x:x['rank']); buyers.sort(key=lambda x:x['rank'])
            return first_contract,sellers[:10],buyers[:10]
    raise ValueError('日経225先物の売超/買超参加者表を特定できません')

def add_week_change(current,previous,contracts_match:bool):
    prev_map={u.norm(x.get('name')):x for x in (previous or [])}
    out=[]
    for item in current or []:
        x=dict(item); p=prev_map.get(u.norm(x.get('name'))) if contracts_match else None
        pv=u.n((p or {}).get('openInterest')); cv=u.n(x.get('openInterest'))
        x['previousOpenInterest']=int(round(pv)) if pv is not None else None
        x['weekChange']=int(round(cv-pv)) if cv is not None and pv is not None else None
        x['previousRank']=(p or {}).get('rank') if p else None
        out.append(x)
    return out

def market_oi_for_date(trade_date:str):
    td=ymd(trade_date)
    url=BASE+f'/markets/derivatives/trading-volume/tvdivq00000014nn-att/{td}open_interest.xlsx'
    parsed=vf.parse_oi(url); large=parsed.get('large')
    if not large or u.n(large.get('openInterest')) is None:
        raise ValueError(f'Nikkei 225 market OI not found for {td}: {parsed}')
    return {**large,'sourceFileUrl':url}

def futures_price_for_date(d:dict,trade_date:str):
    target=str(trade_date)[:10]; foreign=d.get('foreignInvestors') or {}
    candidates=[]
    if foreign.get('asOfDate'): candidates.append(foreign)
    candidates.extend(foreign.get('series') or [])
    for x in candidates:
        if same_day(x.get('asOfDate'),target) and u.n(x.get('nikkeiFuturesPrice')) is not None:
            return {
                'price':u.n(x.get('nikkeiFuturesPrice')),
                'sourceName':foreign.get('futuresPriceSourceName') or '日経225先物時系列',
                'sourceUrl':x.get('futuresPriceSourceUrl') or foreign.get('futuresPriceSourceUrl'),
                'definition':foreign.get('futuresPriceDefinition')
            }
    return None

def build_same_date_analysis(d:dict,oi_td:str,oi_url:str,contract:str,sellers:list,buyers:list):
    target_iso=iso_date(oi_td); errors=[]
    # Daily participant turnover aligned to the weekly OI reference date.
    turnover=None
    try:
        vtd,vurl=participant_volume_file_for_date(oi_td); vcontract,leaders=parse_turnover_leaders(vurl)
        turnover={'asOfDate':iso_date(vtd),'contract':vcontract,'leaders':leaders,'sourceFileUrl':vurl,'status':'verified'}
        if turnover['asOfDate']!=target_iso: errors.append('取引高の基準日不一致')
    except Exception as exc:
        errors.append(f'同日取引高取得失敗: {type(exc).__name__}: {exc}')

    # Previous weekly participant table for week-over-week changes when the same
    # participant remains inside JPX's published ranking.
    prev_td,prev_url=previous_participant_oi_file(oi_td)
    prev_contract=None; prev_sellers=[]; prev_buyers=[]
    if prev_td and prev_url:
        try:
            prev_contract,prev_sellers,prev_buyers=parse_oi_leaders(prev_url)
        except Exception as exc:
            errors.append(f'前週参加者建玉取得失敗: {type(exc).__name__}: {exc}')
    contracts_match=bool(prev_contract and prev_contract==contract)
    enriched_sellers=add_week_change(sellers,prev_sellers,contracts_match)
    enriched_buyers=add_week_change(buyers,prev_buyers,contracts_match)

    # Market-wide OI aligned to the same weekly date and the previous weekly date.
    market=None
    try:
        cur_oi=market_oi_for_date(oi_td)
        prev_oi=market_oi_for_date(prev_td) if prev_td else None
        price_now=futures_price_for_date(d,target_iso)
        price_prev=futures_price_for_date(d,iso_date(prev_td)) if prev_td else None
        if not price_now:
            errors.append('同日先物価格取得不能')
        oi_now=u.n(cur_oi.get('openInterest')); oi_prev=u.n((prev_oi or {}).get('openInterest'))
        p_now=u.n((price_now or {}).get('price')); p_prev=u.n((price_prev or {}).get('price'))
        market={
            'asOfDate':target_iso,
            'previousAsOfDate':iso_date(prev_td) if prev_td else None,
            'price':p_now,
            'previousWeekPrice':p_prev,
            'priceWeekChange':(p_now-p_prev) if p_now is not None and p_prev is not None else None,
            'priceWeekChangePercent':((p_now/p_prev)-1)*100 if p_now is not None and p_prev not in (None,0) else None,
            'openInterest':int(round(oi_now)) if oi_now is not None else None,
            'previousWeekOpenInterest':int(round(oi_prev)) if oi_prev is not None else None,
            'openInterestWeekChange':int(round(oi_now-oi_prev)) if oi_now is not None and oi_prev is not None else None,
            'dailyOpenInterestChange':cur_oi.get('openInterestChange'),
            'volume':cur_oi.get('volume'),
            'oiSourceFileUrl':cur_oi.get('sourceFileUrl'),
            'previousOiSourceFileUrl':(prev_oi or {}).get('sourceFileUrl'),
            'priceSourceName':(price_now or {}).get('sourceName'),
            'priceSourceUrl':(price_now or {}).get('sourceUrl'),
            'priceDefinition':(price_now or {}).get('definition')
        }
    except Exception as exc:
        errors.append(f'同日市場全体建玉取得失敗: {type(exc).__name__}: {exc}')

    verified=bool(turnover and market and u.n(market.get('openInterest')) is not None and u.n(market.get('price')) is not None)
    return {
        'sourceName':'JPX 同一基準日 取引参加者需給比較',
        'comment':'週次の取引参加者別建玉残高の基準日に、日次取引高・日経225先物価格・市場全体建玉を合わせて比較します。異なる基準日のデータは分析判定に混在させません。',
        'asOfDate':target_iso,
        'previousAsOfDate':iso_date(prev_td) if prev_td else None,
        'contract':contract,
        'previousContract':prev_contract,
        'contractsMatch':contracts_match,
        'turnover':turnover,
        'openInterest':{
            'buyers':enriched_buyers,
            'sellers':enriched_sellers,
            'sourceFileUrl':oi_url,
            'previousSourceFileUrl':prev_url,
            'status':'verified'
        },
        'market':market,
        'status':'verified' if verified and not errors else 'partial' if verified or turnover or market else 'unavailable',
        'error':' / '.join(errors) if errors else None,
        'fetchedAt':u.now()
    }

def main():
    d=json.loads(OUT.read_text(encoding='utf-8'))
    prev=d.get('participantFlow') or {}
    base={'sourceName':'JPX 取引参加者別取引高（手口上位一覧）','sourceUrl':'https://www.jpx.co.jp/markets/derivatives/participant-volume/','comment':'日次ファイルは売買方向別ではなく取引高上位です。取引高上位を売り・買いと読み替えません。'}
    try:
        td,url=latest_participant_volume_file(); contract,leaders=parse_turnover_leaders(url)
        d['participantFlow']={**base,'asOfDate':iso_date(td),'contract':contract,'leaders':leaders,'sourceFileUrl':url,'status':'verified','fetchedAt':u.now()}
    except Exception as exc:
        d['participantFlow']=u.stale(prev,base,f'JPX手口上位取得失敗: {type(exc).__name__}: {exc}')

    prev=d.get('participantOpenInterest') or {}
    base={'sourceName':'JPX 取引参加者別建玉残高','sourceUrl':'https://www.jpx.co.jp/markets/derivatives/open-interest/','comment':'週次の期近限月について、JPXが明示する売超参加者・買超参加者の上位を表示します。最終投資家を直接示すものではありません。'}
    current_oi=None
    try:
        td,url=latest_participant_oi_file(); contract,sellers,buyers=parse_oi_leaders(url)
        d['participantOpenInterest']={**base,'asOfDate':iso_date(td),'contract':contract,'sellers':sellers,'buyers':buyers,'sourceFileUrl':url,'status':'verified','fetchedAt':u.now()}
        current_oi=(td,url,contract,sellers,buyers)
    except Exception as exc:
        d['participantOpenInterest']=u.stale(prev,base,f'JPX参加者別建玉取得失敗: {type(exc).__name__}: {exc}')

    if current_oi:
        td,url,contract,sellers,buyers=current_oi
        d['sameDateParticipantAnalysis']=build_same_date_analysis(d,td,url,contract,sellers,buyers)
    else:
        old=d.get('sameDateParticipantAnalysis') or {}
        d['sameDateParticipantAnalysis']=u.stale(old,{'sourceName':'JPX 同一基準日 取引参加者需給比較'},'週次参加者建玉が取得できないため同一基準日分析を更新できません')

    keys=('spot','futures','sessions','arbitrage','options','participantFlow','foreignInvestors','participantOpenInterest','shortSelling','margin')
    statuses={k:(d.get(k) or {}).get('status','unavailable') for k in keys}
    connected=sum(v in {'verified','calculated'} for v in statuses.values())
    d['sourceStatus']=f'{connected}/10主要セクション稼働（各内部指標の欠損は個別表示）'
    d.setdefault('diagnostics',{})['statuses']=statuses
    d['diagnostics']['participantParser']='latest daily turnover + latest weekly participant OI + same-date turnover/price/market-OI comparison anchored to weekly OI date'
    d['diagnostics']['sameDateParticipantAnalysis']=(d.get('sameDateParticipantAnalysis') or {}).get('status','unavailable')
    d['generatedAt']=u.now()
    OUT.write_text(json.dumps(d,ensure_ascii=False,indent=2)+'\n',encoding='utf-8')
    print(json.dumps({'participantFlow':d['participantFlow'],'participantOpenInterest':d['participantOpenInterest'],'sameDateParticipantAnalysis':d['sameDateParticipantAnalysis'],'sourceStatus':d['sourceStatus']},ensure_ascii=False))
if __name__=='__main__': main()
