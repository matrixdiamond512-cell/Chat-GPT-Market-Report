#!/usr/bin/env python3
from __future__ import annotations
import io,json,re,sys
from pathlib import Path
from openpyxl import load_workbook
ROOT=Path(__file__).resolve().parents[1]; sys.path.insert(0,str(ROOT/'scripts'))
import update_nikkei225_supply_demand as u
OUT=ROOT/'data/nikkei225-session-row-inspection.json'
links=u.links(u.URLS['futures'])
urls={}
for url,_ in links:
    if 'derivatives_market_data_whole_day.xlsx' in url: urls['wholeDay']=url
    elif 'derivatives_market_data_night.xlsx' in url: urls['night']=url
out={'generatedAt':u.now(),'sources':{}}
for key,url in urls.items():
    wb=load_workbook(io.BytesIO(u.get(url).content),read_only=True,data_only=True)
    sheets=[]
    for ws in wb.worksheets:
        rows=list(ws.iter_rows(values_only=True))
        matches=[]
        for i,row in enumerate(rows):
            text=' | '.join(u.txt(x) for x in row if x not in (None,''))
            if re.search(r'日経\s*225|Nikkei\s*225',text,re.I):
                lo=max(0,i-3); hi=min(len(rows),i+8)
                for j in range(lo,hi):
                    cells=[[c,u.txt(v)] for c,v in enumerate(rows[j]) if v not in (None,'')]
                    matches.append({'row':j+1,'cells':cells})
        if matches:
            # deduplicate row numbers
            seen=set(); uniq=[]
            for x in matches:
                if x['row'] not in seen: seen.add(x['row']); uniq.append(x)
            sheets.append({'title':ws.title,'rows':uniq[:120]})
    out['sources'][key]={'url':url,'sheets':sheets}
OUT.write_text(json.dumps(out,ensure_ascii=False,indent=2)+'\n',encoding='utf-8')
print(OUT)
