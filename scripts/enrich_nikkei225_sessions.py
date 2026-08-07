#!/usr/bin/env python3
"""Add verified day/night participation data for Nikkei 225 futures.

JPX's current same-day overview publishes session-specific trading volume. It does
not provide a directly comparable session price-change field in the same file,
so this module exposes verified volumes instead of inventing price changes.
"""
from __future__ import annotations
import io,json,re
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
import update_nikkei225_supply_demand as u
ROOT=Path(__file__).resolve().parents[1]; OUT=ROOT/'data/nikkei225-supply-demand.json'

def latest_night_url():
    xs=[url for url,_ in u.links(u.URLS['futures']) if re.search(r'derivatives_market_data_night\.xlsx(?:\?|$)',url,re.I)]
    return xs[0] if xs else None

def night_large_volume(url:str):
    wb=load_workbook(io.BytesIO(u.get(url).content),read_only=True,data_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            vals=list(row)
            for i,v in enumerate(vals):
                s=u.txt(v)
                if re.fullmatch(r'日経225先物(?:\s+Nikkei\s+225\s+Futures)?',s,re.I):
                    # Current market_data_Futures layout: product, session,
                    # total volume, J-NET volume, total value, J-NET value.
                    session=u.txt(vals[i+1]) if i+1<len(vals) else ''
                    volume=u.n(vals[i+2]) if i+2<len(vals) else None
                    if '夜間' in session or 'Night' in session:
                        return int(round(volume)) if volume is not None else None
    return None

def main():
    d=json.loads(OUT.read_text(encoding='utf-8')); prev=d.get('sessions') or {}; f=d.get('futures') or {}
    base={'sourceName':'JPX デリバティブ取引市況（ナイト・日通し）','sourceUrl':u.URLS['futures'],'comment':'JPX公式のセッション別出来高を使用。価格差が同一基準で確認できないため、日中/ナイトの価格変化は推測しません。'}
    try:
        url=latest_night_url()
        if not url: raise ValueError('night-session workbook link not found')
        night=night_large_volume(url); total=u.n(f.get('volume'))
        if night is None or total is None: raise ValueError(f'night={night}, total={total}')
        day=int(round(total))-night
        if day<0: raise ValueError(f'total volume smaller than night volume: total={total}, night={night}')
        denom=day+night
        m=re.search(r'(20\d{6})',url); asof=datetime.strptime(m.group(1),'%Y%m%d').date().isoformat() if m else f.get('asOfDate')
        d['sessions']={**base,'dayChange':None,'nightChange':None,'dayVolume':day,'nightVolume':night,'daySharePercent':day/denom*100 if denom else None,'nightSharePercent':night/denom*100 if denom else None,'dayDriver':'日中出来高','nightDriver':'ナイト出来高','asOfDate':asof,'nightSourceFileUrl':url,'status':'verified','fetchedAt':u.now()}
    except Exception as exc:
        d['sessions']=u.stale(prev,base,f'JPXセッション別出来高取得失敗: {type(exc).__name__}: {exc}')
    keys=('spot','futures','sessions','arbitrage','options','participantFlow','foreignInvestors','participantOpenInterest','shortSelling','margin'); statuses={k:(d.get(k) or {}).get('status','unavailable') for k in keys}; d['sourceStatus']=f"{sum(v in {'verified','calculated'} for v in statuses.values())}/10項目連携（基準日を個別表示）"; d.setdefault('diagnostics',{})['statuses']=statuses; d['diagnostics']['sessionParser']='JPX night volume + verified whole-day Nikkei 225 futures volume'; d['generatedAt']=u.now(); OUT.write_text(json.dumps(d,ensure_ascii=False,indent=2)+'\n',encoding='utf-8'); print(json.dumps(d['sessions'],ensure_ascii=False))
if __name__=='__main__':main()
