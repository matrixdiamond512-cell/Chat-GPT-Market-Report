#!/usr/bin/env python3
"""Complete Nikkei participant analysis with same-date price and market-wide OI.

The weekly participant-OI date is the anchor. Historical market-wide Nikkei 225
futures OI comes from JPX Monthly Statistics, Index Futures Trading (Daily).
For dates whose monthly file is not published yet, verified daily observations
saved by this updater are used instead.
"""
from __future__ import annotations
import io,json
from datetime import datetime,date
from pathlib import Path
from openpyxl import load_workbook
import update_nikkei225_supply_demand as u

ROOT=Path(__file__).resolve().parents[1]
OUT=ROOT/'data/nikkei225-supply-demand.json'
BASE='https://www.jpx.co.jp'

def iso(v): return str(v or '')[:10]
def parse_date(v): return datetime.strptime(iso(v),'%Y-%m-%d').date()

def remember_current(d):
    hist=d.get('nikkeiFuturesDailyHistory') or []
    by={iso(x.get('asOfDate')):dict(x) for x in hist if iso(x.get('asOfDate'))}
    f=d.get('futures') or {}
    day=iso(f.get('asOfDate'))
    if day and u.n(f.get('openInterest')) is not None and u.n(f.get('price')) is not None:
        by[day]={
            'asOfDate':day,
            'price':u.n(f.get('price')),
            'openInterest':int(round(u.n(f.get('openInterest')))),
            'volume':int(round(u.n(f.get('volume')))) if u.n(f.get('volume')) is not None else None,
            'sourceName':f.get('sourceName') or 'JPX / 大阪取引所 日経225先物',
            'priceSourceUrl':f.get('sourceUrl'),
            'oiSourceFileUrl':f.get('oiSourceUrl'),
            'capturedAt':u.now(),
            'status':'verified'
        }
    rows=sorted(by.values(),key=lambda x:iso(x.get('asOfDate')),reverse=True)[:160]
    d['nikkeiFuturesDailyHistory']=rows
    return rows

def hist_record(d,day):
    target=iso(day)
    for x in d.get('nikkeiFuturesDailyHistory') or []:
        if iso(x.get('asOfDate'))==target and u.n(x.get('openInterest')) is not None:
            return dict(x)
    return None

def monthly_sif_record(day):
    dt=parse_date(day); ym=dt.strftime('%Y%m')
    url=BASE+f'/automation/markets/statistics-derivatives/monthly-statistics/files/{dt.year}/SIF_D_{ym}.xlsx'
    wb=load_workbook(io.BytesIO(u.get(url).content),read_only=True,data_only=True)
    for ws in wb.worksheets:
        for raw in ws.iter_rows(values_only=True):
            r=list(raw)
            if len(r)<11: continue
            if u.txt(r[1])!='日経225先物' and u.txt(r[2])!='Nikkei 225 Futures': continue
            token=u.txt(r[0]).replace('月','.')
            if not token: continue
            try:
                n=float(token)
                row_day=int(round((n-int(n))*10)) if n!=int(n) else int(n)
                if '.' in token: row_day=int(token.split('.')[-1])
            except Exception:
                try: row_day=int(float(token))
                except Exception: continue
            if row_day!=dt.day: continue
            oi=u.n(r[10]); vol=u.n(r[4])
            if oi is None: continue
            return {
                'asOfDate':dt.isoformat(),
                'openInterest':int(round(oi)),
                'volume':int(round(vol)) if vol is not None else None,
                'sourceName':'JPX 月間統計資料「指数先物取引取引状況（日別）」',
                'sourceFileUrl':url,
                'status':'verified'
            }
    raise ValueError(f'Nikkei 225 Futures row not found for {dt.isoformat()} in {url}')

def market_record(d,day):
    # Prefer official monthly archive when it has already been published.
    try: return monthly_sif_record(day)
    except Exception as monthly_error:
        h=hist_record(d,day)
        if h:
            h['archiveFallbackReason']=f'{type(monthly_error).__name__}: {monthly_error}'
            return h
        raise monthly_error

def price_record(d,day):
    target=iso(day)
    # Current/saved daily price first.
    for x in d.get('nikkeiFuturesDailyHistory') or []:
        if iso(x.get('asOfDate'))==target and u.n(x.get('price')) is not None:
            return {'price':u.n(x.get('price')),'sourceName':x.get('sourceName'),'sourceUrl':x.get('priceSourceUrl'),'definition':'日次保存値'}
    foreign=d.get('foreignInvestors') or {}
    candidates=[]
    if foreign.get('asOfDate'): candidates.append(foreign)
    candidates.extend(foreign.get('series') or [])
    for x in candidates:
        if iso(x.get('asOfDate'))==target and u.n(x.get('nikkeiFuturesPrice')) is not None:
            return {
                'price':u.n(x.get('nikkeiFuturesPrice')),
                'sourceName':foreign.get('futuresPriceSourceName') or '日経225先物時系列',
                'sourceUrl':x.get('futuresPriceSourceUrl') or foreign.get('futuresPriceSourceUrl'),
                'definition':foreign.get('futuresPriceDefinition')
            }
    return None

def clean_errors(text):
    if not text:return []
    keep=[]
    for part in str(text).split(' / '):
        if part.startswith('同日市場全体建玉取得失敗:'): continue
        if part=='同日先物価格取得不能': continue
        keep.append(part)
    return keep

def main():
    d=json.loads(OUT.read_text(encoding='utf-8'))
    remember_current(d)
    comp=d.get('sameDateParticipantAnalysis') or {}
    anchor=iso(comp.get('asOfDate')); previous=iso(comp.get('previousAsOfDate'))
    errors=clean_errors(comp.get('error'))
    market=None
    if anchor:
        try:
            cur=market_record(d,anchor)
            prev=market_record(d,previous) if previous else None
            pn=price_record(d,anchor); pp=price_record(d,previous) if previous else None
            if not pn: errors.append('同日先物価格取得不能')
            if previous and not pp: errors.append('前週先物価格取得不能')
            oi=u.n(cur.get('openInterest')); poi=u.n((prev or {}).get('openInterest'))
            p=u.n((pn or {}).get('price')); pprev=u.n((pp or {}).get('price'))
            market={
                'asOfDate':anchor,
                'previousAsOfDate':previous or None,
                'price':p,
                'previousWeekPrice':pprev,
                'priceWeekChange':p-pprev if p is not None and pprev is not None else None,
                'priceWeekChangePercent':((p/pprev)-1)*100 if p is not None and pprev not in (None,0) else None,
                'openInterest':int(round(oi)) if oi is not None else None,
                'previousWeekOpenInterest':int(round(poi)) if poi is not None else None,
                'openInterestWeekChange':int(round(oi-poi)) if oi is not None and poi is not None else None,
                'volume':cur.get('volume'),
                'oiSourceName':cur.get('sourceName'),
                'oiSourceFileUrl':cur.get('sourceFileUrl') or cur.get('oiSourceFileUrl'),
                'previousOiSourceFileUrl':(prev or {}).get('sourceFileUrl') or (prev or {}).get('oiSourceFileUrl'),
                'priceSourceName':(pn or {}).get('sourceName'),
                'priceSourceUrl':(pn or {}).get('sourceUrl'),
                'priceDefinition':(pn or {}).get('definition'),
                'status':'verified' if oi is not None and p is not None else 'partial'
            }
        except Exception as exc:
            errors.append(f'同日市場全体建玉取得失敗: {type(exc).__name__}: {exc}')
    if market is not None: comp['market']=market
    ready=bool(
        comp.get('turnover') and comp.get('openInterest') and market and
        u.n(market.get('price')) is not None and u.n(market.get('previousWeekPrice')) is not None and
        u.n(market.get('openInterest')) is not None and u.n(market.get('previousWeekOpenInterest')) is not None
    )
    comp['status']='verified' if ready and not errors else ('partial' if comp.get('turnover') or market else 'unavailable')
    comp['error']=' / '.join(errors) if errors else None
    comp['marketAlignment']='weekly participant OI anchor + same-date JPX SIF_D market OI + same-date futures price'
    comp['fetchedAt']=u.now()
    d['sameDateParticipantAnalysis']=comp
    d.setdefault('diagnostics',{})['sameDateParticipantAnalysis']=comp.get('status')
    d['diagnostics']['sameDateMarketOiParser']='JPX monthly-statistics SIF_D_YYYYMM.xlsx; current-month fallback to retained verified daily history'
    d['generatedAt']=u.now()
    OUT.write_text(json.dumps(d,ensure_ascii=False,indent=2)+'\n',encoding='utf-8')
    print(json.dumps({'status':comp.get('status'),'asOfDate':anchor,'market':market,'error':comp.get('error')},ensure_ascii=False))
if __name__=='__main__': main()
