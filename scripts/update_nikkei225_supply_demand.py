#!/usr/bin/env python3
from __future__ import annotations
import csv, io, json, math, re
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from urllib.parse import urljoin
import holidays, requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from pypdf import PdfReader
ROOT=Path(__file__).resolve().parents[1]; OUT=ROOT/'data/nikkei225-supply-demand.json'; MARKET=ROOT/'data/market/latest.json'; STOCKS=ROOT/'data/stocks.json'
JST=timezone(timedelta(hours=9)); UA='Mozilla/5.0 (compatible; ChatGPT-Market-Report/1.0)'
URLS={'futures':'https://www.jpx.co.jp/markets/derivatives/trading-volume/','arbitrage':'https://www.jpx.co.jp/markets/statistics-equities/program/','options':'https://www.jpx.co.jp/markets/derivatives/option-price/01.html','flow':'https://www.jpx.co.jp/markets/derivatives/participant-volume/','cash':'https://www.jpx.co.jp/markets/statistics-equities/investor-type/','sector':'https://www.jpx.co.jp/markets/statistics-derivatives/sector/','oi':'https://www.jpx.co.jp/markets/derivatives/open-interest/','short':'https://www.jpx.co.jp/markets/statistics-equities/short-selling/','margin':'https://www.jpx.co.jp/markets/statistics-equities/margin/03.html'}
LARGE=re.compile(r'(?:日経\s*225\s*先物|Nikkei\s*225\s*Futures)',re.I); MINI=re.compile(r'(?:日経\s*225\s*mini|Nikkei\s*225\s*mini)',re.I); TOPIX=re.compile(r'TOPIX.*(?:先物|Futures)',re.I); FOREIGN=re.compile(r'(?:海外投資家|外国人|Foreign\s+Investors?|Overseas)',re.I)
def now(): return datetime.now(JST).replace(microsecond=0).isoformat()
def load(p,d):
 try:return json.loads(p.read_text(encoding='utf-8'))
 except:return d
def n(v):
 try:
  x=float(str(v).replace(',','').strip()); return x if math.isfinite(x) else None
 except:return None
def txt(v): return re.sub(r'\s+',' ',str(v or '').strip())
def norm(v): return re.sub(r'[\s　_\-/（）()・:：]','',txt(v)).lower()
def get(url):
 r=requests.get(url,headers={'User-Agent':UA,'Accept':'*/*'},timeout=35); r.raise_for_status(); return r
def stale(prev,base,err):
 if isinstance(prev,dict) and any(v not in (None,'',[],{}) for k,v in prev.items() if k not in {'status','error','sourceName','sourceUrl','comment','fetchedAt'}):
  x=dict(prev); x.update(base); x['status']='stale'; x['error']=err; x['fetchedAt']=now(); return x
 return {**base,'status':'unavailable','error':err,'fetchedAt':now()}
def links(page):
 s=BeautifulSoup(get(page).text,'html.parser'); out=[]
 for a in s.find_all('a',href=True):
  u=urljoin(page,a['href'])
  if re.search(r'\.(csv|xlsx|pdf)(\?|$)',u,re.I): out.append((u,txt(a.parent.get_text(' ',strip=True) if a.parent else a.get_text(' ',strip=True))))
 return out[:40]
def decode(b):
 for e in ('utf-8-sig','cp932','shift_jis','utf-8'):
  try:return b.decode(e)
  except UnicodeDecodeError:pass
 return b.decode('utf-8',errors='replace')
def doc(url):
 b=get(url).content; low=url.lower().split('?')[0]; rows=[]; text=''
 if low.endswith('.pdf'):
  text='\n'.join((p.extract_text() or '') for p in PdfReader(io.BytesIO(b)).pages); rows=[[x] for x in text.splitlines()]
 elif low.endswith('.xlsx'):
  wb=load_workbook(io.BytesIO(b),read_only=True,data_only=True)
  for ws in wb.worksheets: rows += [list(r) for r in ws.iter_rows(values_only=True)]
  text='\n'.join('\t'.join(txt(x) for x in r if x is not None) for r in rows)
 else:
  text=decode(b); rows=[list(r) for r in csv.reader(io.StringIO(text))]
 return rows,text
def docs(page,keys=(),limit=8):
 ll=links(page); kk=[norm(k) for k in keys]; chosen=[x for x in ll if not kk or any(k in norm(x[1]) for k in kk)] or ll; out=[]
 for u,l in chosen:
  try: out.append((u,l,*doc(u)))
  except: pass
  if len(out)>=limit: break
 return out
def rtext(r): return ' | '.join(txt(x) for x in r if x not in (None,''))
def report_date(ds):
 for _,label,_,text in ds:
  m=re.search(r'(20\d{2})[年/.-](\d{1,2})[月/.-](\d{1,2})',label+' '+text[:1200])
  if m:
   try:return date(*map(int,m.groups())).isoformat()
   except:pass
 return None
def hmap(rows,i):
 best={}
 for j in range(max(0,i-7),i):
  cur={norm(v):k for k,v in enumerate(rows[j]) if txt(v)}
  if len(cur)>len(best):best=cur
 return best
def metric(ds,product,words,exclude=None):
 ww=[norm(x) for x in words]
 for u,_,rows,_ in ds:
  for i,r in enumerate(rows):
   t=rtext(r)
   if not product.search(t) or (exclude and exclude.search(t)):continue
   for h,c in hmap(rows,i).items():
    if c<len(r) and any(w in h for w in ww):
     x=n(r[c])
     if x is not None:return x,u
 return None,None
def labelled(text,labels):
 for q in labels:
  m=re.search(re.escape(q)+r'.{0,35}?([+-]?[\d,]+(?:\.\d+)?)',text,re.I)
  if m and n(m.group(1)) is not None:return n(m.group(1))
 return None
def spot(stocks,prev):
 b={'sourceName':'株式市場分析データ','sourceUrl':'data/stocks.json'}
 try:
  j=stocks['marketInternals']['japan']; r=next(x for x in j['rows'] if isinstance(x,list) and txt(x[0])=='日経225'); v=n(r[1]); d=j.get('dataDate')
  if v is None or not d:raise ValueError('日経225現物または基準日なし')
  return {**b,'value':v,'asOfDate':str(d)[:10],'status':'verified','fetchedAt':now()}
 except Exception as e:return stale(prev,b,f'{type(e).__name__}: {e}')
def futures(market,prev):
 b={'sourceName':'JPX / 大阪取引所 日経225先物','sourceUrl':URLS['futures']}; p=prev or {}; out=dict(b); f=market.get('markets',{}).get('nikkei225_futures_ose',{})
 if f.get('verificationStatus')=='verified' and n(f.get('value')) is not None: out.update(price=n(f['value']),priceChange=n(f.get('change')),priceChangePercent=n(f.get('changePercent')),asOfDate=str(f.get('asOf') or '')[:10] or None,priceAsOf=f.get('asOf'))
 else: out.update({k:p.get(k) for k in ('price','priceChange','priceChangePercent','asOfDate','priceAsOf')})
 errs=[]
 try: ds=docs(URLS['futures'],('日通し','建玉残高','デリバティブ'),10)
 except Exception as e:ds=[];errs.append(str(e))
 for key,product,words,ex in [('volume',LARGE,('取引高','出来高','volume'),MINI),('openInterest',LARGE,('建玉残高','建玉','openinterest','oi'),MINI),('miniVolume',MINI,('取引高','出来高','volume'),None),('miniOpenInterest',MINI,('建玉残高','建玉','openinterest','oi'),None)]:
  v,_=metric(ds,product,words,ex); out[key]=int(round(v)) if v is not None else p.get(key)
  if v is None and key in ('volume','openInterest'): errs.append(key+'列を安全に特定できません')
 d=report_date(ds); oldd=p.get('asOfDate'); newday=bool(d and d!=oldd)
 out['openInterestChange']=(n(out.get('openInterest'))-n(p.get('openInterest'))) if newday and n(out.get('openInterest')) is not None and n(p.get('openInterest')) is not None else None
 out['miniOpenInterestChange']=(n(out.get('miniOpenInterest'))-n(p.get('miniOpenInterest'))) if newday and n(out.get('miniOpenInterest')) is not None and n(p.get('miniOpenInterest')) is not None else None
 out['volumeChangePercent']=(n(out['volume'])/n(p['volume'])-1)*100 if newday and n(out.get('volume')) is not None and n(p.get('volume')) not in (None,0) else None
 out['miniVolumeChangePercent']=(n(out['miniVolume'])/n(p['miniVolume'])-1)*100 if newday and n(out.get('miniVolume')) is not None and n(p.get('miniVolume')) not in (None,0) else None
 if d:out['asOfDate']=d
 out['status']='verified' if n(out.get('price')) is not None and not errs else 'partial' if n(out.get('price')) is not None else 'unavailable'; out['fetchedAt']=now()
 if errs:out['error']=' / '.join(errs)
 return out
def simple_source(prev,key,comment,labels):
 b={'sourceName':labels[0],'sourceUrl':URLS[key],'comment':comment}
 try:
  ds=docs(URLS[key],labels[1:],8); text='\n'.join(x[3] for x in ds); return b,ds,text
 except Exception as e: return None,None,stale(prev,b,f'{type(e).__name__}: {e}')
def arbitrage(prev):
 b,ds,x=simple_source(prev,'arbitrage','裁定買い残の増減から現物への買い波及を確認。前々営業日データとして表示。',('JPX 裁定取引の状況','裁定'))
 if b is None:return x
 buy=labelled(x,('裁定買い残','買いポジション','買残高','買い残高')); sell=labelled(x,('裁定売り残','売りポジション','売残高','売り残高'))
 if buy is None and sell is None:return stale(prev,b,'裁定買い残・売り残のラベル付き数値を特定できません')
 d=report_date(ds); p=prev or {}; ch=bool(d and d!=p.get('asOfDate'))
 return {**b,'buyBalance':buy,'buyChange':buy-n(p.get('buyBalance')) if ch and buy is not None and n(p.get('buyBalance')) is not None else None,'sellBalance':sell,'sellChange':sell-n(p.get('sellBalance')) if ch and sell is not None and n(p.get('sellBalance')) is not None else None,'asOfDate':d,'status':'verified','fetchedAt':now()}
def sq_options(prev):
 b={'sourceName':'JPX 日経225オプション','sourceUrl':URLS['options'],'comment':'Put/Call・IV・SQ接近を分け、オプションだけで価格目標を断定しません。'}; t=datetime.now(JST).date(); d=date(t.year,t.month,1)
 while d.weekday()!=4:d+=timedelta(days=1)
 d+=timedelta(days=7)
 if d<t:
  y,m=(t.year+1,1) if t.month==12 else (t.year,t.month+1); d=date(y,m,1)
  while d.weekday()!=4:d+=timedelta(days=1)
  d+=timedelta(days=7)
 jp=holidays.Japan(years={t.year,d.year}); bd=sum(1 for k in range(1,(d-t).days+1) if (t+timedelta(days=k)).weekday()<5 and t+timedelta(days=k) not in jp)
 return {**b,'nextSqDate':d.isoformat(),'businessDaysToSq':bd,'putCallRatio':None,'iv':None,'ivChange':None,'asOfDate':t.isoformat(),'status':'calculated','error':'Put/Call・IVはJPX列構造を確認できるまで推測せず空欄','fetchedAt':now()}
def rank_docs(page,prev,oi=False):
 b={'sourceName':'JPX 取引参加者別建玉残高' if oi else 'JPX 取引参加者別取引高（手口上位一覧）','sourceUrl':page,'comment':'手口は最終投資家を直接示すものではありません。' if not oi else '上位参加者への集中度を確認します。'}
 try: ds=docs(page,('指数先物','日中','立会','xlsx','csv'),8)
 except Exception as e:return stale(prev,b,str(e))
 out={'buyers':[],'sellers':[]}
 for side in ('buyers','sellers'):
  sw=('買','buy') if side=='buyers' else ('売','sell'); vals=[]
  for _,_,rows,_ in ds:
   for i,r in enumerate(rows):
    t=rtext(r)
    if not LARGE.search(t) or MINI.search(t):continue
    hm=hmap(rows,i); name=None; qty=None
    for h,c in hm.items():
     if c>=len(r):continue
     if any(k in h for k in ('取引参加者','participant','証券会社','会社名')):name=txt(r[c])
     if any(k in h for k in sw):qty=n(r[c])
    if name and qty is not None:vals.append({'name':name,('openInterest' if oi else 'volume'):int(round(qty))})
  q='openInterest' if oi else 'volume'; vals.sort(key=lambda z:z[q],reverse=True); seen=set()
  for z in vals:
   if z['name'] not in seen:out[side].append(z);seen.add(z['name'])
   if len(out[side])==3:break
 if not out['buyers'] and not out['sellers']:return stale(prev,b,'日経225先物の買い/売り列を安全に特定できません')
 return {**b,**out,'asOfDate':report_date(ds),'status':'verified','fetchedAt':now()}
def net_sector(ds,prod):
 for _,_,rows,_ in ds:
  for i,r in enumerate(rows):
   t=rtext(r)
   if not prod.search(t) or not FOREIGN.search(t):continue
   hm=hmap(rows,i); buy=sell=netv=None
   for h,c in hm.items():
    if c>=len(r):continue
    if '差引' in h or 'net' in h:netv=n(r[c])
    elif '買' in h or 'buy' in h:buy=n(r[c])
    elif '売' in h or 'sell' in h:sell=n(r[c])
   if netv is not None:return netv
   if buy is not None and sell is not None:return buy-sell
 return None
def foreign(prev):
 b={'sourceName':'JPX 投資部門別売買状況','sourceUrl':URLS['sector'],'cashNote':'現物の方向を確認','nikkeiNote':'ヘッジ・短期売買を確認','topixNote':'大型株全体との整合性を確認','comment':'週次データを当日の主体と断定せず、現物と先物の組み合わせを読みます。'}
 try: dd=docs(URLS['sector'],('csv','Tousi','自己委託'),4); cd=docs(URLS['cash'],('金額','xlsx','株式週間'),4)
 except Exception as e:return stale(prev,b,str(e))
 nik=net_sector(dd,LARGE); top=net_sector(dd,TOPIX); cash=net_sector(cd,re.compile(r'.*'))
 if cash is None and nik is None and top is None:return stale(prev,b,'海外投資家の現物・先物差引を安全に特定できません')
 return {**b,'cashNet':cash,'nikkeiFuturesNet':nik,'topixFuturesNet':top,'asOfDate':report_date(dd) or report_date(cd),'status':'verified','fetchedAt':now()}
def short_margin(prev,key):
 if key=='short':b={'sourceName':'JPX 空売り集計','sourceUrl':URLS[key],'comment':'空売り比率は5日・20日平均と比較します。'}; labs=('空売り集計',)
 else:b={'sourceName':'JPX 信用取引現在高','sourceUrl':URLS[key],'comment':'信用需給は日本株個人投資家の補助指標です。'};labs=('信用取引現在高','xlsx','csv')
 try:ds=docs(URLS[key],labs,22 if key=='short' else 4); text='\n'.join(x[3] for x in ds)
 except Exception as e:return stale(prev,b,str(e))
 if key=='short':
  r=labelled(text,('空売り比率','Short Selling Ratio'))
  if r is None:return stale(prev,b,'空売り比率を安全に特定できません')
  return {**b,'ratio':r,'avg5':None,'avg20':None,'asOfDate':report_date(ds),'status':'verified','fetchedAt':now()}
 buy=labelled(text,('信用買い残','買残高','買い残高')); sell=labelled(text,('信用売り残','売残高','売り残高'))
 if buy is None and sell is None:return stale(prev,b,'信用買い残・売り残を安全に特定できません')
 return {**b,'buyBalance':buy,'sellBalance':sell,'ratio':buy/sell if buy is not None and sell not in (None,0) else None,'asOfDate':report_date(ds),'status':'verified','fetchedAt':now()}
def assessment(f,a,o,fo):
 pc=n(f.get('priceChange')); oi=n(f.get('openInterestChange')); s='判定待ち'
 if pc is not None and oi is not None:s='強気（新規買い）' if pc>0 and oi>0 else '上昇（買い戻し）' if pc>0 and oi<0 else '弱気（新規売り）' if pc<0 and oi>0 else '下落（手仕舞い）' if pc<0 and oi<0 else '中立'
 ac=n(a.get('buyChange')); ar='強い' if ac is not None and ac>0 else 'やや弱い' if ac is not None and ac<0 else '中立' if ac==0 else '判定待ち'; op='中立（SQ日程のみ）' if o.get('nextSqDate') else '判定待ち'; c=n(fo.get('cashNet')); nf=n(fo.get('nikkeiFuturesNet')); fr='判定待ち'
 if c is not None and nf is not None:fr=('現物買い・先物買い' if c>0 and nf>0 else '現物買い・先物売り' if c>0 and nf<0 else '現物売り・先物買い' if c<0 and nf>0 else '現物売り・先物売り')
 ev=sum(x!='判定待ち' for x in (s,ar,fr)); score=(2 if '強気' in s else 1 if '上昇' in s else -2 if '弱気' in s else -1 if '下落' in s else 0)+(1 if ar=='強い' else -1 if '弱い' in ar else 0)+(2 if fr=='現物買い・先物買い' else -2 if fr=='現物売り・先物売り' else 1 if '現物買い' in fr else -1 if '現物売り' in fr else 0); overall='判定保留' if ev<2 else '買い優勢' if score>=3 else 'やや買い優勢' if score>=1 else '売り優勢' if score<=-3 else 'やや売り優勢' if score<=-1 else '中立'
 return {'overall':overall,'shortTerm':s,'arbitrage':ar,'options':op,'foreign':fr,'reason':f'短期先物={s}、裁定={ar}、オプション={op}、海外投資家={fr}。各データの鮮度を分離して判定。','comment':'価格×建玉を中心に、裁定・オプション・海外投資家を補助線として重ねます。'}
def main():
 p=load(OUT,{}); m=load(MARKET,{}); st=load(STOCKS,{}); d={'schemaVersion':'2.1.0','pageId':'nikkei225-supply-demand','pageTitle':'日経225需給分析','generatedAt':now()}
 d['spot']=spot(st,p.get('spot')); d['futures']=futures(m,p.get('futures')); d['sessions']=stale(p.get('sessions'),{'sourceName':'JPX / 大阪取引所','sourceUrl':URLS['futures'],'comment':'日中とナイトを分けて確認します。'},'セッション始値/終値列を安全に特定できるまで自動算出を保留'); d['arbitrage']=arbitrage(p.get('arbitrage')); d['options']=sq_options(p.get('options')); d['participantFlow']=rank_docs(URLS['flow'],p.get('participantFlow')); d['foreignInvestors']=foreign(p.get('foreignInvestors')); d['participantOpenInterest']=rank_docs(URLS['oi'],p.get('participantOpenInterest'),True); d['shortSelling']=short_margin(p.get('shortSelling'),'short'); d['margin']=short_margin(p.get('margin'),'margin'); d['assessment']=assessment(d['futures'],d['arbitrage'],d['options'],d['foreignInvestors']); d['watchpoints']=['先物建玉の増加継続が本物の上昇か確認','裁定買い残が増えるかどうか','SQ接近でオプション主導の振れに注意','海外投資家の週次データで現物買い継続を確認']; keys=('spot','futures','sessions','arbitrage','options','participantFlow','foreignInvestors','participantOpenInterest','shortSelling','margin'); ss=[d[k].get('status') for k in keys]; d['sourceStatus']=f"{sum(x in ('verified','calculated') for x in ss)}/10項目連携（基準日を個別表示）"; d['diagnostics']={'policy':'primary-source-first-no-fabrication','statuses':dict(zip(keys,ss))}; OUT.write_text(json.dumps(d,ensure_ascii=False,indent=2)+'\n',encoding='utf-8'); print(d['sourceStatus'])
if __name__=='__main__':main()
