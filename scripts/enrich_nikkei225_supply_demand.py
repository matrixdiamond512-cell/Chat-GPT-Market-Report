#!/usr/bin/env python3
"""Enrich Nikkei 225 supply-demand JSON with JPX-format-specific parsers.

This module intentionally handles only structures that have been verified against
current JPX public files. Unknown layouts remain unavailable instead of being
filled by heuristics.
"""
from __future__ import annotations

import io
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import update_nikkei225_supply_demand as u

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "nikkei225-supply-demand.json"


def _number(v: Any) -> float | None:
    return u.n(v)


def _integer(v: Any) -> int | None:
    x = _number(v)
    return int(round(x)) if x is not None else None


def _latest_link(page: str, pattern: str) -> str | None:
    rx = re.compile(pattern, re.I)
    candidates = [url for url, _ in u.links(page) if rx.search(url)]
    return candidates[0] if candidates else None


def _date_from_filename(url: str) -> str | None:
    m = re.search(r"(20\d{6})", url)
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1), "%Y%m%d").date().isoformat()
    except ValueError:
        return None


def _is_contract(v: Any) -> bool:
    return bool(re.search(r"20\d{2}年\d{2}月限", u.txt(v)))


def _read_oi_totals(url: str) -> dict[str, dict[str, int | None]]:
    """Read official JPX OI workbook.

    Current workbook has two horizontal product blocks. Each block has six
    columns for product/contract, volume, current OI, OI change and prior OI.
    A product name is shown only on its first contract row; its total row starts
    with '合計'. We keep state per horizontal block and capture the total row.
    """
    wb = load_workbook(io.BytesIO(u.get(url).content), read_only=True, data_only=True)
    targets = {"日経225": "large", "日経225mini": "mini"}
    found: dict[str, dict[str, int | None]] = {}

    for ws in wb.worksheets:
        states: dict[int, str | None] = {0: None, 6: None}
        sums: dict[tuple[int, str], dict[str, int]] = {}
        for raw in ws.iter_rows(values_only=True):
            row = list(raw)
            for base in (0, 6):
                if base >= len(row):
                    continue
                first = u.txt(row[base])
                first_norm = u.norm(first)

                if first in targets:
                    states[base] = targets[first]
                    sums[(base, states[base])] = {"volume": 0, "openInterest": 0, "openInterestChange": 0, "previousOpenInterest": 0}
                    # first contract row: product, contract, volume, OI, change, prior OI
                    if base + 5 < len(row):
                        bucket = sums[(base, states[base])]
                        for key, idx in (("volume", base + 2), ("openInterest", base + 3), ("openInterestChange", base + 4), ("previousOpenInterest", base + 5)):
                            val = _integer(row[idx])
                            if val is not None:
                                bucket[key] += val
                    continue

                state = states.get(base)
                if state is None:
                    continue

                # Product total rows are shifted one column left because the
                # product-name merged cell is no longer repeated.
                if first_norm in {"合計", "total"}:
                    vals = {
                        "volume": _integer(row[base + 1]) if base + 1 < len(row) else None,
                        "openInterest": _integer(row[base + 2]) if base + 2 < len(row) else None,
                        "openInterestChange": _integer(row[base + 3]) if base + 3 < len(row) else None,
                        "previousOpenInterest": _integer(row[base + 4]) if base + 4 < len(row) else None,
                    }
                    # Fall back to summed contract rows if a total cell is blank.
                    bucket = sums.get((base, state), {})
                    for key in vals:
                        if vals[key] is None and key in bucket:
                            vals[key] = bucket[key]
                    found[state] = vals
                    states[base] = None
                    continue

                # Continuation contract rows normally have a blank product cell
                # and contract month in base+1.
                if not first and base + 1 < len(row) and _is_contract(row[base + 1]):
                    bucket = sums[(base, state)]
                    for key, idx in (("volume", base + 2), ("openInterest", base + 3), ("openInterestChange", base + 4), ("previousOpenInterest", base + 5)):
                        if idx < len(row):
                            val = _integer(row[idx])
                            if val is not None:
                                bucket[key] += val
                    continue

                # If a different named product starts in this block, stop
                # attributing subsequent rows to the previous product.
                if first and not _is_contract(first) and first_norm not in {"合計", "total"}:
                    states[base] = None

    return found


def enrich_futures(d: dict[str, Any]) -> None:
    current = d.get("futures") or {}
    base = {
        "sourceName": "JPX / 大阪取引所 日経225先物",
        "sourceUrl": u.URLS["futures"],
    }
    try:
        url = _latest_link(u.URLS["futures"], r"open_interest\.xlsx(?:\?|$)")
        if not url:
            raise ValueError("JPX open_interest.xlsx link not found")
        totals = _read_oi_totals(url)
        large, mini = totals.get("large"), totals.get("mini")
        if not large or large.get("openInterest") is None:
            raise ValueError("Nikkei 225 total OI row not found")

        current.update(base)
        current["volume"] = large.get("volume")
        current["openInterest"] = large.get("openInterest")
        current["openInterestChange"] = large.get("openInterestChange")
        if mini:
            current["miniVolume"] = mini.get("volume")
            current["miniOpenInterest"] = mini.get("openInterest")
            current["miniOpenInterestChange"] = mini.get("openInterestChange")
        asof = _date_from_filename(url)
        if asof:
            current["asOfDate"] = asof
        current["oiSourceUrl"] = url
        current["status"] = "verified" if _number(current.get("price")) is not None else "partial"
        current.pop("error", None)
        current["fetchedAt"] = u.now()
    except Exception as exc:
        current.update(base)
        current["status"] = "partial" if _number(current.get("price")) is not None else "unavailable"
        current["error"] = f"JPX建玉取得失敗: {type(exc).__name__}: {exc}"
        current["fetchedAt"] = u.now()
    d["futures"] = current


def _parse_arbitrage_pdf(url: str) -> dict[str, Any] | None:
    _, text = u.doc(url)
    # The official PDF contains a compact all-participants position line:
    # 株 数 [sell-next] [sell-total] [buy-current] [buy-next] [buy-total]
    m = re.search(
        r"株\s*数\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)",
        text,
        re.S,
    )
    if not m:
        return None
    nums = [_integer(x) for x in m.groups()]
    if any(x is None for x in nums):
        return None

    asof = None
    dm = re.search(r"現物ポジション[（(](\d{1,2})月(\d{1,2})日現在[）)]", text)
    ym = re.search(r"(20\d{2})年\d{1,2}月\d{1,2}日", text)
    if dm and ym:
        try:
            asof = date(int(ym.group(1)), int(dm.group(1)), int(dm.group(2))).isoformat()
        except ValueError:
            pass
    return {"asOfDate": asof, "sellBalance": nums[1], "buyBalance": nums[4], "sourceFileUrl": url}


def enrich_arbitrage(d: dict[str, Any]) -> None:
    prev = d.get("arbitrage") or {}
    base = {
        "sourceName": "JPX 裁定取引の状況",
        "sourceUrl": u.URLS["arbitrage"],
        "comment": "裁定買い・売りポジションはJPXの全取引参加者報告合計。前々営業日データとして鮮度を分離表示。",
    }
    try:
        pdfs = [url for url, _ in u.links(u.URLS["arbitrage"]) if re.search(r"\.pdf(?:\?|$)", url, re.I)]
        parsed = []
        for url in pdfs[:8]:
            try:
                x = _parse_arbitrage_pdf(url)
                if x and x.get("asOfDate"):
                    parsed.append(x)
            except Exception:
                continue
        parsed.sort(key=lambda x: x["asOfDate"], reverse=True)
        if not parsed:
            raise ValueError("JPX arbitrage position rows not found")
        cur = parsed[0]
        old = parsed[1] if len(parsed) > 1 else None
        out = {
            **base,
            **cur,
            "sellChange": cur["sellBalance"] - old["sellBalance"] if old else None,
            "buyChange": cur["buyBalance"] - old["buyBalance"] if old else None,
            "status": "verified",
            "fetchedAt": u.now(),
        }
        d["arbitrage"] = out
    except Exception as exc:
        d["arbitrage"] = u.stale(prev, base, f"JPX裁定取得失敗: {type(exc).__name__}: {exc}")


def _parse_short_pdf(url: str) -> dict[str, Any] | None:
    _, text = u.doc(url)
    m = re.search(
        r"(20\d{2})年(\d{1,2})月(\d{1,2})日\s+"
        r"[\d,]+\s+([\d.]+)%\s+"
        r"[\d,]+\s+([\d.]+)%\s+"
        r"[\d,]+\s+([\d.]+)%\s+[\d,]+",
        text,
    )
    if not m:
        return None
    y, mo, day = map(int, m.group(1, 2, 3))
    regulated = float(m.group(5))
    exempt = float(m.group(6))
    return {"asOfDate": date(y, mo, day).isoformat(), "ratio": regulated + exempt, "sourceFileUrl": url}


def enrich_short_selling(d: dict[str, Any]) -> None:
    prev = d.get("shortSelling") or {}
    base = {
        "sourceName": "JPX 空売り集計",
        "sourceUrl": u.URLS["short"],
        "comment": "空売り比率は価格規制あり・なしの合計。5日・20日平均と比較し、単独で弱気判定しません。",
    }
    try:
        pdfs = [url for url, _ in u.links(u.URLS["short"]) if re.search(r"\.pdf(?:\?|$)", url, re.I)]
        rows = []
        for url in pdfs[:24]:
            try:
                x = _parse_short_pdf(url)
                if x:
                    rows.append(x)
            except Exception:
                continue
        uniq = {x["asOfDate"]: x for x in rows}
        rows = sorted(uniq.values(), key=lambda x: x["asOfDate"], reverse=True)
        if not rows:
            raise ValueError("short-selling ratio rows not found")
        vals = [float(x["ratio"]) for x in rows]
        cur = rows[0]
        d["shortSelling"] = {
            **base,
            **cur,
            "avg5": sum(vals[:5]) / len(vals[:5]) if vals else None,
            "avg20": sum(vals[:20]) / len(vals[:20]) if vals else None,
            "sampleCount": min(20, len(vals)),
            "status": "verified",
            "fetchedAt": u.now(),
        }
    except Exception as exc:
        d["shortSelling"] = u.stale(prev, base, f"JPX空売り取得失敗: {type(exc).__name__}: {exc}")


def refresh_assessment(d: dict[str, Any]) -> None:
    d["assessment"] = u.assessment(
        d.get("futures") or {},
        d.get("arbitrage") or {},
        d.get("options") or {},
        d.get("foreignInvestors") or {},
    )
    keys = (
        "spot", "futures", "sessions", "arbitrage", "options",
        "participantFlow", "foreignInvestors", "participantOpenInterest",
        "shortSelling", "margin",
    )
    statuses = {k: (d.get(k) or {}).get("status", "unavailable") for k in keys}
    connected = sum(v in {"verified", "calculated"} for v in statuses.values())
    d["sourceStatus"] = f"{connected}/10項目連携（基準日を個別表示）"
    diag = d.setdefault("diagnostics", {})
    diag["policy"] = "primary-source-first-no-fabrication"
    diag["statuses"] = statuses
    diag["formatSpecificParsers"] = ["JPX open_interest.xlsx", "JPX arbitrage PDF", "JPX short-selling PDF"]
    d["generatedAt"] = u.now()


def main() -> int:
    d = json.loads(OUT.read_text(encoding="utf-8"))
    enrich_futures(d)
    enrich_arbitrage(d)
    enrich_short_selling(d)
    refresh_assessment(d)
    OUT.write_text(json.dumps(d, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(d.get("sourceStatus"))
    print(json.dumps(d.get("diagnostics", {}), ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
