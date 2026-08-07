#!/usr/bin/env python3
from __future__ import annotations
import io,json,re,sys
from pathlib import Path
from openpyxl import load_workbook
ROOT=Path(__file__).resolve().parents[1]; sys.path.insert(0,str(ROOT/'scripts'))
import update_nikkei225_supply_demand as u
OUT=ROOT/'data/nikkei225-oi-layout.json'
links=[url for url,_ in u.links(u.URLS['futures']) if re.search(r'open_interest\.xlsx(?:\?|$)',url,re.I)]
if not links: raise SystemExit('open_interest.xlsx not found')
url=links[0]; wb=load_workbook(io.BytesIO(u.get(url).content),read_only=True,data_only=True)
out={'generatedAt':u.now(),'sourceUrl':url,'sheets':[]}
for ws in wb.worksheets:
    rows=[]
    for ri,row in enumerate(ws.iter_rows(values_only=True),1):
        cells=[[ci,str(v)] for ci,v in enumerate(row) if v not in (None,'')]
        if cells: rows.append({'row':ri,'cells':cells})
        if ri>=90: break
    out['sheets'].append({'title':ws.title,'rows':rows})
OUT.write_text(json.dumps(out,ensure_ascii=False,indent=2)+'\n',encoding='utf-8')
print(OUT)
