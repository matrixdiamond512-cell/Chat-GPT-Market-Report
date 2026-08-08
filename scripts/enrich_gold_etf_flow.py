#!/usr/bin/env python3
"""Enrich gold ETF data with chart history and the latest public WGC monthly snapshot.

Rules:
- Never add GLD and IAU changes from different as-of dates.
- Backfill GLD history from the official SPDR archive.
- Accumulate IAU snapshots across workflow runs.
- Join only identical dates into etf.historyDaily.
- Use only public WGC values that can be parsed from an official monthly report.
"""
from __future__ import annotations

import io
import json
import math
import re
from calendar import monthrange
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data" / "gold-supply-demand.json"
HISTORY = ROOT / "data" / "gold-etf-flow-history.json"
JST = timezone(timedelta(hours=9))
UA = "Mozilla/5.0 (compatible; ChatGPT-Market-Report/1.0; +https://github.com/matrixdiamond512-cell/Chat-GPT-Market-Report)"
GLD_XLSX = "https://api.spdrgoldshares.com/api/v1/historical-archive?exchange=NYSE&lang=en&product=gld"
WGC_BASE = "https://www.gold.org/goldhub/research/gold-etfs-holdings-and-flows"


def now_iso() -> str:
    return datetime.now(JST).replace(microsecond=0).isoformat()


def num(v: Any) -> float | None:
    try:
        x = float(str(v).replace(",", "").strip())
        return x if math.isfinite(x) else None
    except Exception:
        return None


def parse_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v or "").strip()
    for f in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d-%b-%Y", "%b %d, %Y", "%B %d, %Y", "%d %B %Y", "%d %b %Y"):
        try:
            return datetime.strptime(s, f).date()
        except ValueError:
            pass
    return None


def request(url: str, timeout: int = 35) -> requests.Response:
    r = requests.get(url, timeout=timeout, headers={"User-Agent": UA, "Accept": "text/html,application/xhtml+xml,application/json,*/*"})
    r.raise_for_status()
    return r


def load_json(path: Path, default: Any) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def dedupe(rows: list[dict[str, Any]], limit: int = 120) -> list[dict[str, Any]]:
    by_date: dict[str, dict[str, Any]] = {}
    for row in rows:
        d = str(row.get("asOfDate") or "")[:10]
        if d:
            by_date[d] = {"asOfDate": d, "tonnes": num(row.get("tonnes")), "changeTonnes": num(row.get("changeTonnes"))}
    return [by_date[k] for k in sorted(by_date)][-limit:]


def append_snapshot(rows: list[dict[str, Any]], item: dict[str, Any]) -> list[dict[str, Any]]:
    d = str(item.get("asOfDate") or "")[:10]
    tonnes = num(item.get("tonnes"))
    change = num(item.get("changeTonnes"))
    if d and tonnes is not None:
        rows.append({"asOfDate": d, "tonnes": tonnes, "changeTonnes": change})
    return dedupe(rows)


def gld_archive_history() -> list[dict[str, Any]]:
    content = request(GLD_XLSX, timeout=45).content
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    found: list[tuple[date, float]] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        header_i = date_col = value_col = None
        for ri, row in enumerate(rows[:80]):
            texts = [str(x or "").strip().lower() for x in row]
            for ci, text in enumerate(texts):
                if date_col is None and (text == "date" or "date" in text):
                    date_col = ci
                if value_col is None and any(k in text for k in ("tonnes", "tonnes of gold", "gold holdings")):
                    value_col = ci
            if date_col is not None and value_col is not None:
                header_i = ri
                break
        if header_i is None or date_col is None or value_col is None:
            continue
        for row in rows[header_i + 1 :]:
            if date_col >= len(row) or value_col >= len(row):
                continue
            d = parse_date(row[date_col])
            v = num(row[value_col])
            if d and v is not None:
                found.append((d, v))
    found = sorted(set(found), key=lambda x: x[0])
    out: list[dict[str, Any]] = []
    prev = None
    for d, tonnes in found[-121:]:
        change = tonnes - prev if prev is not None else None
        out.append({"asOfDate": d.isoformat(), "tonnes": tonnes, "changeTonnes": change})
        prev = tonnes
    return out[-120:]


def month_candidates(today: date) -> list[tuple[int, int]]:
    y, m = today.year, today.month
    out = []
    for _ in range(5):
        out.append((y, m))
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return out


def report_month_end(text: str) -> tuple[str | None, str | None]:
    m = re.search(r"Gold ETF Flows:\s*([A-Za-z]+)\s+(20\d{2})", text, re.I)
    if not m:
        return None, None
    try:
        dt = datetime.strptime(f"{m.group(1)} {m.group(2)}", "%B %Y")
        last = monthrange(dt.year, dt.month)[1]
        return f"{dt.year:04d}-{dt.month:02d}", date(dt.year, dt.month, last).isoformat()
    except Exception:
        return None, None


def signed_change(text: str) -> float | None:
    patterns = [
        (r"holdings[^.]{0,150}?(reduced|fell|declined|decreased|dropped|shed|lost)[^0-9]{0,30}([\d,]+(?:\.\d+)?)\s*t\b", -1),
        (r"holdings[^.]{0,150}?(rose|increased|grew|added|gained|rebounded)[^0-9]{0,30}([\d,]+(?:\.\d+)?)\s*t\b", 1),
    ]
    for pattern, sign in patterns:
        m = re.search(pattern, text, re.I)
        if m:
            value = num(m.group(2))
            if value is not None:
                return sign * value
    return None


def fetch_wgc_global(previous: dict[str, Any] | None) -> dict[str, Any]:
    errors = []
    today = datetime.now(JST).date()
    for year, month in month_candidates(today):
        url = f"{WGC_BASE}/{year}/{month:02d}"
        try:
            html = request(url).text
            text = re.sub(r"\s+", " ", BeautifulSoup(html, "html.parser").get_text(" ", strip=True))
            if "Gold ETF" not in text or "holdings" not in text.lower():
                raise ValueError("not a gold ETF monthly report")
            holdings = None
            for pattern in (
                r"(?:collective|total|global gold ETF)?\s*holdings[^.]{0,260}?\b(?:to|at)\s*([\d,]+(?:\.\d+)?)\s*t\b",
                r"holdings[^.]{0,180}?([34],[\d]{3}(?:\.\d+)?)\s*t\b",
            ):
                m = re.search(pattern, text, re.I)
                if m:
                    holdings = num(m.group(1))
                    if holdings is not None:
                        break
            if holdings is None:
                raise ValueError("global holdings tonnage not found")

            period, inferred_asof = report_month_end(text)
            asof = None
            am = re.search(r"As of\s+(\d{1,2}\s+[A-Za-z]+\s+20\d{2})", text, re.I)
            if am:
                d = parse_date(am.group(1))
                asof = d.isoformat() if d else None
            if not asof:
                asof = inferred_asof

            monthly_change = signed_change(text)
            ytd_change = None
            ym = re.search(r"holdings[^.]{0,220}?(?:first half|H1|year[- ]to[- ]date|y[- ]t[- ]d)[^.]{0,150}?(?:up|rose|increased|grew)[^0-9]{0,25}([\d,]+(?:\.\d+)?)\s*t\b", text, re.I)
            if ym:
                ytd_change = num(ym.group(1))

            flow_usd = None
            fm = re.search(r"(?:recorded|saw|posted)?\s*(?:modest\s+)?(inflows|outflows)\s+of\s+US\$([\d,.]+)\s*bn", text, re.I)
            if fm:
                flow_usd = num(fm.group(2))
                if flow_usd is not None and fm.group(1).lower().startswith("out"):
                    flow_usd *= -1

            return {
                "status": "verified",
                "asOfDate": asof,
                "period": period,
                "tonnes": holdings,
                "changeTonnes": monthly_change,
                "ytdChangeTonnes": ytd_change,
                "monthlyFlowUsdBillion": flow_usd,
                "frequency": "monthly",
                "sourceName": "World Gold Council Gold ETF Flows",
                "sourceUrl": url,
                "fetchedAt": now_iso(),
                "note": "WGC公式月次レポート本文から公開値を抽出。週次チャートがログイン必須の場合は月次公開値を使用。",
            }
        except Exception as exc:
            errors.append(f"{url}: {type(exc).__name__}: {exc}")

    prev = dict(previous) if isinstance(previous, dict) else {}
    if prev and any(num(prev.get(k)) is not None for k in ("tonnes", "changeTonnes", "ytdChangeTonnes")):
        prev["status"] = "stale"
        prev["error"] = "WGC月次取得失敗: " + " | ".join(errors[-2:])
        return prev
    return {"status": "unavailable", "sourceName": "World Gold Council Gold ETF Flows", "sourceUrl": WGC_BASE, "error": "WGC月次取得失敗: " + " | ".join(errors[-2:])}


def main() -> int:
    data = load_json(DATA, {})
    history = load_json(HISTORY, {"schemaVersion": "1.0.0", "gld": [], "iau": []})
    etf = data.setdefault("etf", {})
    gld = etf.get("gld") or {}
    iau = etf.get("iau") or {}

    gld_rows = list(history.get("gld") or [])
    try:
        archive = gld_archive_history()
        if archive:
            gld_rows = archive
    except Exception as exc:
        history["gldError"] = f"{type(exc).__name__}: {exc}"
    gld_rows = append_snapshot(gld_rows, gld)

    iau_rows = append_snapshot(list(history.get("iau") or []), iau)
    history.update({"schemaVersion": "1.0.0", "updatedAt": now_iso(), "gld": gld_rows, "iau": iau_rows})

    gld_by = {r["asOfDate"]: r for r in gld_rows if r.get("asOfDate")}
    iau_by = {r["asOfDate"]: r for r in iau_rows if r.get("asOfDate")}
    aligned = []
    for d in sorted(set(gld_by) & set(iau_by)):
        gr, ir = gld_by[d], iau_by[d]
        gc, ic = num(gr.get("changeTonnes")), num(ir.get("changeTonnes"))
        if gc is None or ic is None:
            continue
        aligned.append({
            "asOfDate": d,
            "gldTonnes": num(gr.get("tonnes")),
            "iauTonnes": num(ir.get("tonnes")),
            "gldChangeTonnes": gc,
            "iauChangeTonnes": ic,
            "combinedChangeTonnes": gc + ic,
            "aligned": True,
        })
    etf["historyDaily"] = aligned[-90:]
    etf["historyDaysCount"] = len(etf["historyDaily"])
    etf["historyStatus"] = "verified" if etf["historyDaily"] else "collecting"
    etf["global"] = fetch_wgc_global(etf.get("global") if isinstance(etf.get("global"), dict) else None)

    HISTORY.write_text(json.dumps(history, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    DATA.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "updatedAt": history["updatedAt"],
        "gldHistory": len(gld_rows),
        "iauHistory": len(iau_rows),
        "alignedHistory": len(etf["historyDaily"]),
        "globalStatus": (etf.get("global") or {}).get("status"),
        "globalAsOf": (etf.get("global") or {}).get("asOfDate"),
    }, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
