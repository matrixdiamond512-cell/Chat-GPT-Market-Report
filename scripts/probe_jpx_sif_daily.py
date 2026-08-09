#!/usr/bin/env python3
import io,json,requests
from pathlib import Path
from openpyxl import load_workbook
URL='https://www.jpx.co.jp/automation/markets/statistics-derivatives/monthly-statistics/files/2026/SIF_D_202607.xlsx'
OUT=Path(__file__).resolve().parents[1]/'data/debug-jpx-sif-daily.json'
h={'User-Agent':'Mozilla/5.0 market-report-maintenance/1.0'}
def main():
 r=requests.get(URL,headers=h,timeout=30); r.raise_for_status()
 wb=load_workbook(io.BytesIO(r.content),read_only=True,data_only=True)
 out={'url':URL,'sheets':[]}
 for ws in wb.worksheets:
  rows=[]
  for raw in ws.iter_rows(values_only=True):
   vals=[str(x).strip() if x is not None else '' for x in raw]
   if not any(vals): continue
   rows.append(vals[:30])
   if len(rows)>=160: break
  out['sheets'].append({'title':ws.title,'rows':rows})
 OUT.write_text(json.dumps(out,ensure_ascii=False,indent=2)+'\n',encoding='utf-8')
 print(json.dumps({'sheets':[x['title'] for x in out['sheets']]},ensure_ascii=False))
if __name__=='__main__':main()
