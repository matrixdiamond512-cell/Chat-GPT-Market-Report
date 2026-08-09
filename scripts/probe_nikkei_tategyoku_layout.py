#!/usr/bin/env python3
from __future__ import annotations
import io,json,re
from pathlib import Path
from urllib.parse import urljoin
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import update_nikkei225_supply_demand as u

ROOT=Path(__file__).resolve().parents[1]
OUT=ROOT/'data/nikkei225-supply-demand.json'
SECTOR='https://www.jpx.co.jp/markets/statistics-derivatives/sector/'
MONTHLY='https://www.jpx.co.jp/markets/statistics-derivatives/monthly-statistics/'
DAILY='https://www.jpx.co.jp/markets/statistics-derivatives/daily/'
PAGES=[SECTOR,SECTOR+'00-archives-00.html',SECTOR+'00-archives-01.html',MONTHLY,DAILY]

def compact(s): return re.sub(r'\s+',' ',str(s or '')).strip()

def crawl_page(page):
 html=u.get(page).text
 soup=BeautifulSoup(html,'html.parser')
 anchors=[]
 for a in soup.find_all('a',href=True):
  href=urljoin(page,a['href']); text=compact(a.get_text(' ',strip=True)); parent=compact(a.parent.get_text(' ',strip=True) if a.parent else '')
  anchors.append({'text':text[:180],'parent':parent[:260],'href':href})
 scripts=[urljoin(page,s.get('src')) for s in soup.find_all('script',src=True)]
 return html,anchors,scripts

def main():
 d=json.loads(OUT.read_text(encoding='utf-8'))
 target=str((d.get('participantOpenInterest') or {}).get('asOfDate') or '').replace('-','')[:8]
 target_month=target[:6]
 tategyoku=[]; archive_pages=[]; monthly_hits=[]; daily_hits=[]; scripts=[]; page_errors=[]
 page_data={}
 for page in PAGES:
  try:
   html,anchors,ss=crawl_page(page); page_data[page]=(html,anchors); scripts.extend(ss)
   for a in anchors:
    href=a['href']; blob=(a['text']+' '+a['parent']+' '+href)
    if '建玉内容集計表' in blob or 'Open Interest by Type' in blob:
     archive_pages.append(a)
    if re.search(r'Tategyoku_W_.*\.xlsx(?:\?|$)',href,re.I): tategyoku.append(href)
    if page.startswith(MONTHLY) and (target_month in blob or '2026' in blob or re.search(r'\.(xlsx|xls|csv|zip)(?:\?|$)',href,re.I)):
     monthly_hits.append(a)
    if page.startswith(DAILY) and (target in blob or target_month in blob or re.search(r'\.(zip|pdf|xlsx|csv)(?:\?|$)',href,re.I)):
     daily_hits.append(a)
  except Exception as exc:
   page_errors.append({'page':page,'error':f'{type(exc).__name__}: {exc}'})

 # Follow any discovered archive/navigation pages inside JPX and look for target files.
 follow=[]
 for a in archive_pages+monthly_hits+daily_hits:
  h=a.get('href','')
  if h.startswith('https://www.jpx.co.jp/') and h.endswith('.html') and h not in PAGES:
   follow.append(h)
 for page in list(dict.fromkeys(follow))[:30]:
  try:
   html,anchors,ss=crawl_page(page); scripts.extend(ss)
   for a in anchors:
    href=a['href']; blob=(a['text']+' '+a['parent']+' '+href)
    if re.search(r'Tategyoku_W_.*\.xlsx(?:\?|$)',href,re.I): tategyoku.append(href)
    if target_month in blob or target in blob or re.search(r'\.(xlsx|xls|csv|zip|pdf)(?:\?|$)',href,re.I):
     if 'monthly-statistics' in page: monthly_hits.append(a)
     if '/daily/' in page: daily_hits.append(a)
  except Exception:
   pass

 tategyoku=list(dict.fromkeys(tategyoku)); monthly_hits=list({x['href']:x for x in monthly_hits}.values()); daily_hits=list({x['href']:x for x in daily_hits}.values())
 chosen=next((x for x in tategyoku if target and target in x.rsplit('/',1)[-1]),None)
 probe={
  'target':target,
  'archivePages':archive_pages[:20],
  'tategyokuLinkCount':len(tategyoku),
  'tategyokuLinks':[x.rsplit('/',1)[-1] for x in tategyoku[:40]],
  'chosenTategyoku':chosen,
  'monthlyHits':monthly_hits[:60],
  'dailyHits':daily_hits[:60],
  'scripts':list(dict.fromkeys(scripts))[:60],
  'pageErrors':page_errors,
 }
 if chosen:
  try:
   wb=load_workbook(io.BytesIO(u.get(chosen).content),read_only=True,data_only=True)
   sheets=[]
   for ws in wb.worksheets:
    rows=[]
    for raw in ws.iter_rows(values_only=True):
     vals=[u.txt(x) if x is not None else '' for x in raw]
     if not any(vals): continue
     rows.append(vals[:22])
     if len(rows)>=100: break
    sheets.append({'title':ws.title,'rows':rows})
   probe['tategyokuSheets']=sheets
  except Exception as exc:
   probe['tategyokuError']=f'{type(exc).__name__}: {exc}'
 d.setdefault('diagnostics',{})['historicalOiProbe']=probe
 d['diagnostics'].pop('tategyokuProbe',None)
 d['generatedAt']=u.now()
 OUT.write_text(json.dumps(d,ensure_ascii=False,indent=2)+'\n',encoding='utf-8')
 print(json.dumps(probe,ensure_ascii=False))
if __name__=='__main__': main()
